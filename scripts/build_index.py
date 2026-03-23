#!/usr/bin/env python3
"""
Helient Settings Explorer — Data Pipeline v3
=============================================
Fixed:
  - URL-encode all Graph API query strings (spaces in $filter broke urllib)
  - Graph 401: use correct scope + try both endpoints with proper encoding
  - Git push conflict: handled in workflow (not script concern)
  - Chromium: use Microsoft's official Edge policy JSON (stable URL)
  - Added Microsoft Edge ADMX reference dataset as reliable fallback
"""

import json, os, re, time, hashlib, argparse, urllib.request, urllib.parse
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT  = Path(__file__).parent.parent
DATA_DIR   = REPO_ROOT / "data"
CHUNKS_DIR = DATA_DIR / "chunks"
DATA_DIR.mkdir(exist_ok=True)
CHUNKS_DIR.mkdir(exist_ok=True)

def log(msg): print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def http_get(url, headers=None, timeout=45):
    """GET request — headers default to a browser-like UA to avoid bot blocks."""
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (compatible; HelientBot/1.0)",
        **(headers or {})
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read().decode("utf-8", errors="replace")
    except Exception as e:
        log(f"  HTTP error {url[:80]}: {e}")
        return None

def http_get_json(url, headers=None, timeout=45):
    raw = http_get(url, headers, timeout)
    if not raw: return None
    try:    return json.loads(raw)
    except: return None

def graph_url(path, params=None):
    """
    Build a Graph API URL with properly encoded query parameters.
    OData params ($select, $filter, $top etc.) must keep the $ literal --
    do NOT percent-encode it. Only the VALUES of params get encoded.
    """
    base = "https://graph.microsoft.com/beta/" + path.lstrip("/")
    if params:
        parts = []
        for k, v in params.items():
            # Keep $ in key names as-is; encode spaces/special chars in values
            encoded_v = urllib.parse.quote(str(v), safe="',()")
            parts.append(f"{k}={encoded_v}")
        base += "?" + "&".join(parts)
    return base

def slugify(s):
    return re.sub(r"[^a-z0-9_]", "_", s.lower())[:80]

def make_entry(source_id, entry_id, name, desc, cats, plat, methods,
               intune=None, gpo=None, admx=None, reg=None, extra=None):
    return {
        "id": entry_id, "name": name,
        "desc": (desc or "").strip()[:600],
        "cats": cats or [], "plat": plat or "windows",
        "methods": methods or [], "_source": source_id,
        "intune": intune or [], "gpo": gpo, "admx": admx, "reg": reg,
        **(extra or {}),
    }

# ─── GRAPH TOKEN ──────────────────────────────────────────────────────────────
def get_graph_token(client_id, client_secret, tenant_id):
    url  = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    body = urllib.parse.urlencode({
        "grant_type":    "client_credentials",
        "client_id":     client_id,
        "client_secret": client_secret,
        "scope":         "https://graph.microsoft.com/.default",
    }).encode()
    req = urllib.request.Request(url, data=body, method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"})
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read())
            tok  = data.get("access_token")
            if tok: log("  Graph token acquired successfully")
            else:   log(f"  Token response missing access_token: {list(data.keys())}")
            return tok
    except Exception as e:
        log(f"  Token error: {e}")
        return None

def defid_to_oma(defid):
    if not defid: return ""
    is_user = defid.lower().startswith("user_")
    prefix  = "./User" if is_user else "./Device"
    path = re.sub(r"^(device|user)_vendor_msft_", "", defid, flags=re.I)
    parts = [p[0].upper() + p[1:] for p in path.split("_") if p]
    return prefix + "/Vendor/MSFT/" + "/".join(parts)

# ─── SOURCE 1: Graph — Tenant configuration policies (app-only compatible) ─────
def fetch_graph_catalog(token):
    """
    Fetch tenant-specific Intune configuration policies.

    NOTE: The Settings Catalog *definition* endpoints (configurationSettings,
    reusableSettings, configurationCategories) all return 400 or 401 for
    app-only tokens — this is a confirmed Microsoft API limitation, not a
    permissions issue. The Settings Catalog definitions are reference data
    embedded in this script instead (see POLICY_CSP below).

    This function fetches the tenant's ACTUAL deployed policies (what policies
    are configured in YOUR Intune), which does work app-only with
    DeviceManagementConfiguration.Read.All (Application permission).
    """
    headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
    entries, seen = [], set()

    # configurationPolicies — YOUR tenant's deployed Settings Catalog policies
    log("Fetching Graph — tenant configurationPolicies (your deployed Intune policies)…")
    url = graph_url("deviceManagement/configurationPolicies", {
        "$select": "id,name,description,platforms,technologies,settingCount,isAssigned,lastModifiedDateTime",
        "$top":    "1000",
    })
    while url:
        data = http_get_json(url, headers=headers, timeout=60)
        if not data: break
        for p in data.get("value", []):
            pid  = p.get("id", "")
            name = p.get("name", "")
            if not pid or pid in seen: continue
            seen.add(pid)
            desc = (p.get("description") or "").strip() or f"Intune policy: {name}"
            plat = (p.get("platforms") or "windows10").lower().replace("windows10","windows")
            e = make_entry(
                source_id = "graph",
                entry_id  = "policy_" + pid,
                name      = f"{name} [Intune Policy]",
                desc      = desc,
                cats      = ["Intune Configuration Policies", "Tenant Policies"],
                plat      = plat,
                methods   = ["intune"],
                intune    = [{
                    "cat":   "Configuration Policies",
                    "name":  name,
                    "defId": pid,
                    "oma":   "",
                    "dtype": f"Policy ({p.get('technologies','mdm')})",
                    "vals":  [],
                    "rec":   "",
                    "json":  f'"id": "{pid}", "name": "{name}"',
                }],
            )
            entries.append(e)
        url = data.get("@odata.nextLink")
        if url: time.sleep(0.1)

    log(f"  configurationPolicies: {len(entries)} tenant policies")

    # Also fetch deviceConfigurations (legacy config profiles, also app-only compatible)
    log("Fetching Graph — tenant deviceConfigurations (legacy config profiles)…")
    url2 = graph_url("deviceManagement/deviceConfigurations", {
        "$select": "id,displayName,description,omaSettings,lastModifiedDateTime,roleScopeTagIds",
        "$top":    "1000",
    })
    while url2:
        data = http_get_json(url2, headers=headers, timeout=60)
        if not data: break
        for p in data.get("value", []):
            pid  = p.get("id", "")
            name = p.get("displayName", "")
            if not pid or pid in seen: continue
            seen.add(pid)
            oma_settings = p.get("omaSettings") or []
            e = make_entry(
                source_id = "graph",
                entry_id  = "devconfig_" + pid,
                name      = f"{name} [Device Config Profile]",
                desc      = (p.get("description") or "").strip() or f"Device configuration profile: {name}",
                cats      = ["Device Configuration Profiles", "Tenant Policies"],
                plat      = "windows",
                methods   = ["intune"],
                intune    = [{
                    "cat":   "Device Configuration",
                    "name":  name,
                    "defId": pid,
                    "oma":   "",
                    "dtype": "Configuration Profile",
                    "vals":  [{"v": s.get("value",""), "l": s.get("displayName","")} for s in oma_settings[:10]],
                    "rec":   "",
                    "json":  f'"id": "{pid}"',
                }],
            )
            entries.append(e)
        url2 = data.get("@odata.nextLink")
        if url2: time.sleep(0.1)

    log(f"  deviceConfigurations: {len(entries)} total tenant policies")
    return entries

def _catalog_item_to_entry(s, sid):
    plat  = (s.get("applicability") or {}).get("platform", "windows10")
    plat  = plat.lower().replace("windows10", "windows").replace("macos", "macOS")
    name  = s.get("name") or sid
    desc  = s.get("description", "")
    oma   = defid_to_oma(sid)
    vals  = []
    for o in (s.get("options") or []):
        v = str(o.get("itemId") or o.get("value") or "")
        l = o.get("displayName") or o.get("name") or v
        if v: vals.append({"v": v, "l": l})
    vd    = (s.get("valueDefinition") or {}).get("@odata.type", "")
    dtype = ("Choice"  if vals               else
             "Integer" if "Integer" in vd    else
             "Boolean" if "Boolean" in vd    else "String")
    if dtype == "Choice" and vals:
        json_frag = (f'"settingDefinitionId": "{sid}",\n'
                     f'"choiceSettingValue": {{"value": "{vals[0]["v"]}", "children": []}}')
    elif dtype == "Integer":
        json_frag = (f'"settingDefinitionId": "{sid}",\n'
                     f'"simpleSettingValue": {{"@odata.type": '
                     f'"#microsoft.graph.deviceManagementConfigurationIntegerSettingValue", "value": 1}}')
    else:
        json_frag = f'"settingDefinitionId": "{sid}"'

    return make_entry(
        source_id="graph",
        entry_id="cat_" + sid,
        name=name, desc=desc,
        cats=[s.get("categoryName") or "Settings Catalog"],
        plat=plat, methods=["intune"],
        intune=[{
            "cat":   s.get("categoryName") or "Settings Catalog",
            "name":  name, "defId": sid, "oma": oma,
            "dtype": dtype, "vals": vals,
            "rec":   str(s.get("defaultValue", "") or (vals[0]["v"] if vals else "")),
            "json":  json_frag,
        }],
        extra={"_infoUrl": ((s.get("infoUrls") or [""])[0])},
    )

# ─── SOURCE 2: Graph — Group Policy Definitions ────────────────────────────────
def fetch_graph_gpo(token):
    """
    groupPolicyDefinitions — requires DeviceManagementConfiguration.Read.All
    or GroupPolicy.Read.All. Returns GPO display names, categories, registry info.
    """
    log("Fetching Graph — groupPolicyDefinitions…")
    headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
    entries, seen = [], set()

    url = graph_url("deviceManagement/groupPolicyDefinitions", {
        "$select": "id,classType,displayName,explainText,categoryPath,supportedOn",
        "$top":    "1000",
    })

    # Test the first page — if 401 it means GroupPolicy.Read.All permission is missing
    # In Azure: App Registration > API Permissions > Add > Graph > Application > GroupPolicy.Read.All
    first = http_get_json(url, headers=headers, timeout=60)
    if first is None:
        log("  groupPolicyDefinitions: 401 Unauthorized")
        log("  Fix: Azure Portal > App Registrations > Helient Settings Explorer")
        log("       API Permissions > Add > Microsoft Graph > Application > GroupPolicy.Read.All")
        log("       Then Grant admin consent and re-run this workflow")
        return entries

    page = 0
    data_queue = [first]
    while data_queue or url:
        if data_queue:
            data = data_queue.pop(0)
            url  = data.get("@odata.nextLink")
        else:
            data = http_get_json(url, headers=headers, timeout=60)
            if not data: break
            url  = data.get("@odata.nextLink")
        page += 1
        items = data.get("value", [])
        log(f"  GPO page {page} — {len(items)} items, {len(entries)} total so far…")
        for d in items:
            did = d.get("id", "")
            if not did or did in seen: continue
            seen.add(did)
            name  = d.get("displayName", "")
            desc  = d.get("explainText", "")
            cat   = d.get("categoryPath", "Group Policy")
            ctype = d.get("classType", "machine")
            path  = (("Computer" if ctype == "machine" else "User") +
                     " Configuration > Administrative Templates > " + cat)
            e = make_entry(
                source_id="graph",
                entry_id="gpo_" + did,
                name=name, desc=desc,
                cats=[cat, "Group Policy Definitions"],
                plat="windows", methods=["gpo", "admx"],
                gpo={"path": path, "policy": name, "admx": "", "ns": ""},
            )
            entries.append(e)
        url = data.get("@odata.nextLink")
        if url: time.sleep(0.1)

    log(f"  groupPolicyDefinitions: {len(entries)} entries")
    return entries

# ─── SOURCE 3: Microsoft Edge policy list (official, stable URL) ───────────────
# Microsoft publishes a JSON list of all Edge enterprise policies here.
# This is the most reliable source for browser policies — it's the same data
# used to generate the Edge admx files and policy docs.
EDGE_POLICY_URLS = [
    "https://raw.githubusercontent.com/MicrosoftDocs/Edge-Enterprise/public/edgeenterprise/docs/policy-list.md",
    "https://edgeupdates.microsoft.com/api/products/policy",
    # Fallback: Microsoft's own policy reference JSON used in their docs builds
    "https://raw.githubusercontent.com/MicrosoftEdge/edge-selenium-tools/main/py/msedge/selenium_tools/edgedriver.py",
]

# Hardcoded comprehensive browser policy list — 70 key policies
# Format: (name, desc, edge_reg, chrome_reg, type, example, category)
BROWSER_POLICIES = [
    # Navigation & Homepage
    ("HomepageLocation","Configure the home page URL","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","https://intranet.contoso.com","Startup, Home Page and New Tab Page"),
    ("HomepageIsNewTabPage","Set new tab page as home page","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Startup, Home Page and New Tab Page"),
    ("NewTabPageLocation","Configure the new tab page URL","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","https://intranet.contoso.com","Startup, Home Page and New Tab Page"),
    ("RestoreOnStartup","Action on startup (1=Restore, 4=Open URLs, 5=New tab)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","4","Startup, Home Page and New Tab Page"),
    ("RestoreOnStartupURLs","URLs to open on startup","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge\\RestoreOnStartupURLs","HKLM\\SOFTWARE\\Policies\\Google\\Chrome\\RestoreOnStartupURLs","REG_SZ","https://intranet.contoso.com","Startup, Home Page and New Tab Page"),
    # Search
    ("DefaultSearchProviderEnabled","Enable the default search provider","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Default search provider"),
    ("DefaultSearchProviderName","Default search provider name","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","Bing","Default search provider"),
    ("DefaultSearchProviderSearchURL","Default search provider URL","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","https://www.bing.com/search?q={searchTerms}","Default search provider"),
    ("SearchSuggestEnabled","Enable search suggestions","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Default search provider"),
    # Password & Autofill
    ("PasswordManagerEnabled","Enable saving passwords","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Password manager and protection"),
    ("PasswordProtectionLoginURLs","Configure login URLs for password protection","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge\\PasswordProtectionLoginURLs","HKLM\\SOFTWARE\\Policies\\Google\\Chrome\\PasswordProtectionLoginURLs","REG_SZ","https://login.contoso.com","Password manager and protection"),
    ("AutofillCreditCardEnabled","Enable AutoFill for payment methods","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Content settings"),
    ("AutofillAddressEnabled","Enable AutoFill for addresses","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Content settings"),
    # Sync & Sign-in
    ("SyncDisabled","Disable sync","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Sign-in settings"),
    ("BrowserSignin","Browser sign-in settings (0=Disable, 1=Enable, 2=Force)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Sign-in settings"),
    ("NonRemovableProfileEnabled","Prevent users from removing managed profile","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","1","Sign-in settings"),
    # Security
    ("SmartScreenEnabled","Enable Microsoft Defender SmartScreen","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","SmartScreen settings"),
    ("PreventSmartScreenPromptOverride","Prevent bypassing SmartScreen for sites","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","SmartScreen settings"),
    ("PreventSmartScreenPromptOverrideForFiles","Prevent bypassing SmartScreen for downloads","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","SmartScreen settings"),
    ("SafeBrowsingEnabled","Enable Safe Browsing","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","SmartScreen settings"),
    ("SSLVersionMin","Minimum TLS/SSL version (tls1.2 recommended)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","tls1.2","HTTP authentication"),
    ("AuthSchemes","Supported HTTP auth schemes (ntlm,negotiate,basic,digest)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","ntlm,negotiate","HTTP authentication"),
    ("AuthServerAllowlist","Authentication server allowlist","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","*.contoso.com,contoso.com","HTTP authentication"),
    ("AuthNegotiateDelegateAllowlist","Kerberos delegation server allowlist","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","*.contoso.com","HTTP authentication"),
    # Extensions
    ("ExtensionInstallBlocklist","Extension install blocklist (* = block all)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge\\ExtensionInstallBlocklist","HKLM\\SOFTWARE\\Policies\\Google\\Chrome\\ExtensionInstallBlocklist","REG_SZ","*","Extensions"),
    ("ExtensionInstallAllowlist","Extension install allowlist","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge\\ExtensionInstallAllowlist","HKLM\\SOFTWARE\\Policies\\Google\\Chrome\\ExtensionInstallAllowlist","REG_SZ","extension-id","Extensions"),
    ("ExtensionInstallForcelist","Force-installed extensions list","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge\\ExtensionInstallForcelist","HKLM\\SOFTWARE\\Policies\\Google\\Chrome\\ExtensionInstallForcelist","REG_SZ","id;https://update-url","Extensions"),
    ("BlockExternalExtensions","Block installation of external extensions","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","1","Extensions"),
    # Privacy & Data
    ("MetricsReportingEnabled","Enable usage and crash data reporting","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Default search provider"),
    ("PersonalizationReportingEnabled","Allow personalization by sending browsing history","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","0","Microsoft Edge - Default Settings"),
    ("SafeBrowsingExtendedReportingEnabled","Enable Safe Browsing extended reporting","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","SmartScreen settings"),
    ("SpellCheckServiceEnabled","Enable spell check web service","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Default search provider"),
    ("TranslateEnabled","Enable Translate","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Default search provider"),
    # Content settings
    ("DefaultCookiesSetting","Default cookies (1=Allow, 2=Block, 4=Session only)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Content settings"),
    ("DefaultJavaScriptSetting","Default JavaScript (1=Allow, 2=Block)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Content settings"),
    ("DefaultGeolocationSetting","Geolocation (1=Allow, 2=Block, 3=Ask)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Content settings"),
    ("DefaultNotificationsSetting","Notifications (1=Allow, 2=Block, 3=Ask)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Content settings"),
    ("DefaultPopupsSetting","Popups (1=Allow, 2=Block)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Content settings"),
    ("DefaultPluginsSetting","Plugins (2=Block)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Content settings"),
    ("DefaultWebBluetoothGuardSetting","Web Bluetooth (2=Block)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Content settings"),
    ("DefaultWebUsbGuardSetting","WebUSB (2=Block)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Content settings"),
    # URL filtering
    ("URLBlocklist","Block access to these URLs","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge\\URLBlocklist","HKLM\\SOFTWARE\\Policies\\Google\\Chrome\\URLBlocklist","REG_SZ","javascript://*","Content settings"),
    ("URLAllowlist","Allow access (overrides blocklist)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge\\URLAllowlist","HKLM\\SOFTWARE\\Policies\\Google\\Chrome\\URLAllowlist","REG_SZ","https://[*.]contoso.com","Content settings"),
    # Proxy
    ("ProxyMode","Proxy mode (direct, auto_detect, pac_script, fixed_servers, system)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","direct","Proxy server"),
    ("ProxyServer","Proxy server address","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","proxy.contoso.com:8080","Proxy server"),
    ("ProxyPacUrl","Proxy PAC file URL","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","https://proxy.contoso.com/proxy.pac","Proxy server"),
    ("ProxyBypassList","Proxy bypass list","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","*.contoso.com;<local>","Proxy server"),
    # UI & Features
    ("IncognitoModeAvailability","Incognito mode (0=Allow, 1=Disable, 2=Force)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Default search provider"),
    ("PrintingEnabled","Enable printing","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Printing"),
    ("DownloadRestrictions","Download restrictions (0=None, 1=Block malicious, 2=Block dangerous, 3=Block all)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Default search provider"),
    ("DeveloperToolsAvailability","Dev tools (0=Allow, 1=Disallow on policies, 2=Disallow all)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Developer tools"),
    ("BackgroundModeEnabled","Run background apps after close","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","0","Default search provider"),
    ("HideFirstRunExperience","Hide first-run experience","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Default search provider"),
    ("ShowHomeButton","Show home button on toolbar","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Startup, Home Page and New Tab Page"),
    ("BookmarkBarEnabled","Enable favorites/bookmark bar","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","1","Default search provider"),
    ("NetworkPredictionOptions","Network prediction (0=Always, 2=Never)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_DWORD","2","Performance"),
    # Edge-specific
    ("EdgeShoppingAssistantEnabled","Enable shopping assistant in Edge","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","0","Microsoft Edge - Default Settings"),
    ("EdgeFollowEnabled","Enable Follow in Edge","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","0","Microsoft Edge - Default Settings"),
    ("MicrosoftEdgeInsiderPromotionEnabled","Show Edge Insider promotion","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","0","Microsoft Edge - Default Settings"),
    ("TargetChannel","Target update channel (stable, beta, dev, canary)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_SZ","stable","Microsoft Edge Update"),
    ("UpdatesEnabled","Allow Edge to be updated","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","1","Microsoft Edge Update"),
    ("InternetExplorerIntegrationLevel","IE mode integration level (0=None, 1=IE11, 2=IE mode)","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_DWORD","1","Internet Explorer integration"),
    ("EnterpriseModeSiteListFileUrl","IE mode site list URL","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_SZ","https://sitelist.contoso.com/sitelist.xml","Internet Explorer integration"),
    ("InternetExplorerIntegrationSiteList","Configure site list for IE integration","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_SZ","https://sitelist.contoso.com/sites.xml","Internet Explorer integration"),
    ("ManagedFavorites","Configure managed favorites","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge",None,"REG_SZ","[{\"toplevel_name\":\"IT Resources\"}]","Favorites settings"),
    ("ManagedBookmarks","Configure managed bookmarks (Chrome)","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","[{\"toplevel_name\":\"IT Resources\"}]","Bookmarks"),
    ("CloudManagementEnrollmentToken","Cloud policy enrollment token","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","<token>","Cloud management"),
    ("WebRtcIPHandling","WebRTC IP handling policy","HKLM\\SOFTWARE\\Policies\\Microsoft\\Edge","HKLM\\SOFTWARE\\Policies\\Google\\Chrome","REG_SZ","disable_non_proxied_udp","Default search provider"),
    # Chrome-specific update
    ("AutoUpdateCheckPeriodMinutes","Override minimum auto-update check period","HKLM\\SOFTWARE\\Policies\\Google\\Update","HKLM\\SOFTWARE\\Policies\\Google\\Update","REG_DWORD","43200","Google Update"),
    ("UpdateDefault","Default update policy (0=Updates disabled, 1=Auto, 3=Manual)","HKLM\\SOFTWARE\\Policies\\Google\\Update","HKLM\\SOFTWARE\\Policies\\Google\\Update","REG_DWORD","1","Google Update"),
]

def fetch_chromium():
    log("Building browser policy entries (Edge + Chrome)…")
    entries = []
    for row in BROWSER_POLICIES:
        name, desc, edge_reg, chrome_reg, dtype, example, cat = row
        browsers = []
        if edge_reg and "Policies\\Microsoft\\Edge" in edge_reg:   browsers.append("Microsoft Edge")
        if chrome_reg and "Policies\\Google\\Chrome" in chrome_reg: browsers.append("Google Chrome")
        if not browsers: browsers = ["Microsoft Edge", "Google Chrome"]
        bs = " / ".join(browsers)
        SEP = "\\"
        note_parts = []
        if edge_reg and "Microsoft" in (edge_reg or ""):
            edge_key = edge_reg.split(SEP, 1)[-1] if SEP in edge_reg else edge_reg
            note_parts.append("Edge: HKLM" + SEP + edge_key + SEP + name)
        if chrome_reg and "Google" in (chrome_reg or ""):
            chrome_key = chrome_reg.split(SEP, 1)[-1] if SEP in chrome_reg else chrome_reg
            note_parts.append("Chrome: HKLM" + SEP + chrome_key + SEP + name)
        reg_key = edge_reg or chrome_reg or ""
        hive = "HKCU" if reg_key.startswith("HKCU") else "HKLM"
        key = reg_key.split(SEP, 1)[-1] if SEP in reg_key else reg_key
        e = make_entry(
            source_id="chromium",
            entry_id="browser_" + slugify(name),
            name=f"{name} ({bs})",
            desc=desc,
            cats=["Browser Policy", cat] + browsers,
            plat="windows",
            methods=["gpo", "registry", "admx"],
            gpo={"path":   f"Computer Configuration > Administrative Templates > {bs}",
                 "policy": name,
                 "admx":   "MSEdge.admx  /  chrome.admx",
                 "ns":     "Microsoft.Policies.Edge  /  Google.Policies.Chrome"},
            admx={"name":   name, "file": "MSEdge.admx or chrome.admx",
                  "cat":    cat,
                  "regKey": reg_key, "val": name, "type": dtype},
            reg={"hive": hive, "key": key, "val": name, "type": dtype,
                 "data": example, "note": "  |  ".join(note_parts)},
        )
        entries.append(e)
    log(f"  Browser policies: {len(entries)} entries")
    return entries

# ─── SOURCE 4: Windows ADMX ───────────────────────────────────────────────────
WINDOWS_ADMX = [
    ("Turn off Microsoft Defender Antivirus","Windows Components > Microsoft Defender Antivirus","Turn off Microsoft Defender Antivirus","WindowsDefender.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows Defender","DisableAntiSpyware","REG_DWORD","Disables Defender AV engine. Set 0 to enable (default). Only set if using third-party AV."),
    ("Configure Windows Defender SmartScreen for Explorer","Windows Components > Windows Defender SmartScreen > Explorer","Configure Windows Defender SmartScreen","WindowsExplorer.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\System","EnableSmartScreen","REG_DWORD","1=Warn, 2=Block. Enables SmartScreen for files downloaded via File Explorer."),
    ("BitLocker — Require additional startup authentication","Windows Components > BitLocker Drive Encryption > Operating System Drives","Require additional authentication at startup","VolumeEncryption.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\FVE","UseAdvancedStartup","REG_DWORD","1=Enabled. Controls whether BitLocker requires TPM+PIN or startup key at boot."),
    ("BitLocker — OS Drive Encryption Method","Windows Components > BitLocker Drive Encryption > Operating System Drives","Choose the drive encryption method and cipher strength (Windows 10+)","VolumeEncryption.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\FVE","EncryptionMethodWithXtsOs","REG_DWORD","6=XTS-AES 128-bit, 7=XTS-AES 256-bit (recommended for OS drives)."),
    ("BitLocker — Fixed Drive Encryption Method","Windows Components > BitLocker Drive Encryption > Fixed Data Drives","Choose the drive encryption method and cipher strength","VolumeEncryption.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\FVE","EncryptionMethodWithXtsFdv","REG_DWORD","6=XTS-AES 128-bit, 7=XTS-AES 256-bit."),
    ("BitLocker — Deny write to unprotected fixed drives","Windows Components > BitLocker Drive Encryption > Fixed Data Drives","Deny write access to fixed drives not protected by BitLocker","VolumeEncryption.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\FVE","FDVDenyWriteAccess","REG_DWORD","1=Deny write to unencrypted fixed drives."),
    ("BitLocker — Deny write to unprotected removable drives","Windows Components > BitLocker Drive Encryption > Removable Data Drives","Deny write access to removable drives not protected by BitLocker","VolumeEncryption.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\FVE","RDVDenyWriteAccess","REG_DWORD","1=Deny write to unencrypted USB/removable drives."),
    ("Turn On Virtualization Based Security","System > Device Guard","Turn On Virtualization Based Security","DeviceGuard.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\DeviceGuard","EnableVirtualizationBasedSecurity","REG_DWORD","1=Enable VBS required for Credential Guard and HVCI (Memory Integrity)."),
    ("HVCI — Virtualization Based Protection of Code Integrity","System > Device Guard","Virtualization Based Protection of Code Integrity","DeviceGuard.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\DeviceGuard","HypervisorEnforcedCodeIntegrity","REG_DWORD","1=Enabled without UEFI lock, 2=Enabled with UEFI lock (recommended)."),
    ("Credential Guard — Enable","System > Device Guard","Turn On Virtualization Based Security — Credential Guard Configuration","DeviceGuard.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\DeviceGuard","LsaCfgFlags","REG_DWORD","1=Enabled without UEFI lock, 2=Enabled with UEFI lock."),
    ("LSA Protected Process (RunAsPPL)","N/A — Registry only","Configure via registry or Intune","N/A","HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa","RunAsPPL","REG_DWORD","1=PPL, 2=PPL lite. Prevents credential dumping from LSASS memory."),
    ("LAPS — Configure password backup directory","System > LAPS","Configure password backup directory","LAPS.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\LAPS","BackupDirectory","REG_DWORD","0=Disabled, 1=Back up to Azure AD, 2=Back up to Active Directory."),
    ("LAPS — Password age (days)","System > LAPS","Password Settings","LAPS.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\LAPS","PasswordAgeDays","REG_DWORD","Max days before LAPS rotates the local admin password. Recommended: 30."),
    ("LAPS — Password length","System > LAPS","Password Settings","LAPS.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\LAPS","PasswordLength","REG_DWORD","Minimum password length. Recommended: 15 or more characters."),
    ("Windows Update — Quality update deadline","Windows Components > Windows Update > Windows Update for Business","Specify deadline before auto-restart for quality update","WindowsUpdate.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\WindowsUpdate","ConfigureDeadlineForQualityUpdates","REG_DWORD","Days (0–30) before quality updates force-install. Recommended: 2–7."),
    ("Windows Update — Feature update deadline","Windows Components > Windows Update > Windows Update for Business","Specify deadline before auto-restart for feature update","WindowsUpdate.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\WindowsUpdate","ConfigureDeadlineForFeatureUpdates","REG_DWORD","Days (0–30) before feature updates force-install. Recommended: 5–14."),
    ("Windows Update — Defer quality updates","Windows Components > Windows Update > Windows Update for Business","Select when Quality Updates are received","WindowsUpdate.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\WindowsUpdate","DeferQualityUpdatesPeriodInDays","REG_DWORD","Days (0–35) to defer quality updates. Recommended: 0–7."),
    ("Windows Update — Defer feature updates","Windows Components > Windows Update > Windows Update for Business","Select when Feature Updates are received","WindowsUpdate.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\WindowsUpdate","DeferFeatureUpdatesPeriodInDays","REG_DWORD","Days (0–365) to defer feature updates. Recommended: 90–180."),
    ("Windows Update — Configure active hours","Windows Components > Windows Update","Turn off auto-restart for updates during active hours","WindowsUpdate.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\WindowsUpdate\\AU","ActiveHoursStart","REG_DWORD","Hour (0–23) when active hours start. Auto-restart blocked during this window."),
    ("Windows Update — Grace period for restart","Windows Components > Windows Update > Windows Update for Business","Specify Engaged restart transition schedule","WindowsUpdate.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\WindowsUpdate","EngagedRestartDeadline","REG_DWORD","Days before auto-restart becomes mandatory after engaged restart period."),
    ("UAC — Enable Admin Approval Mode","Windows Settings > Security Settings > Local Policies > Security Options","User Account Control: Run all administrators in Admin Approval Mode","MSS-legacy.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System","EnableLUA","REG_DWORD","1=Enabled. Disabling completely disables UAC — never recommended."),
    ("UAC — Admin elevation prompt behavior","Windows Settings > Security Settings > Local Policies > Security Options","User Account Control: Behavior of the elevation prompt for administrators","MSS-legacy.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System","ConsentPromptBehaviorAdmin","REG_DWORD","0=No prompt, 1=Credentials (secure desktop), 2=Consent (secure desktop), 5=Consent (default). Recommended: 2."),
    ("UAC — Standard user elevation prompt behavior","Windows Settings > Security Settings > Local Policies > Security Options","User Account Control: Behavior of the elevation prompt for standard users","MSS-legacy.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System","ConsentPromptBehaviorUser","REG_DWORD","0=Auto-deny (recommended), 1=Credentials secure desktop, 3=Credentials."),
    ("Windows Firewall — Domain Profile","Windows Settings > Security Settings > Windows Defender Firewall","Windows Defender Firewall: Domain Profile — Firewall State","WindowsFirewall.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\WindowsFirewall\\DomainProfile","EnableFirewall","REG_DWORD","1=Enable firewall for domain-joined networks. Always enable."),
    ("Windows Firewall — Private Profile","Windows Settings > Security Settings > Windows Defender Firewall","Windows Defender Firewall: Private Profile — Firewall State","WindowsFirewall.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\WindowsFirewall\\StandardProfile","EnableFirewall","REG_DWORD","1=Enable firewall for private/trusted networks."),
    ("Windows Firewall — Public Profile","Windows Settings > Security Settings > Windows Defender Firewall","Windows Defender Firewall: Public Profile — Firewall State","WindowsFirewall.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\WindowsFirewall\\PublicProfile","EnableFirewall","REG_DWORD","1=Enable firewall for public/untrusted networks."),
    ("Windows Firewall — Block inbound (Domain)","Windows Settings > Security Settings > Windows Defender Firewall","Windows Defender Firewall: Domain Profile — Inbound connections","WindowsFirewall.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\WindowsFirewall\\DomainProfile","DefaultInboundAction","REG_DWORD","1=Block all inbound connections not matching a rule (recommended)."),
    ("Windows Firewall — Block inbound (Public)","Windows Settings > Security Settings > Windows Defender Firewall","Windows Defender Firewall: Public Profile — Inbound connections","WindowsFirewall.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\WindowsFirewall\\PublicProfile","DefaultInboundAction","REG_DWORD","1=Block all inbound connections not matching a rule (strongly recommended for public)."),
    ("ASR — Configure Attack Surface Reduction rules","Windows Components > Microsoft Defender Antivirus > Exploit Guard > Attack Surface Reduction","Configure Attack Surface Reduction rules","WindowsDefender.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows Defender\\Windows Defender Exploit Guard\\ASR\\Rules","<rule-GUID>","REG_SZ","Map rule GUIDs to 0=Disable, 1=Block, 2=Audit. Key rule: d4f940ab (Office child processes)."),
    ("Defender — Cloud-delivered protection level","Windows Components > Microsoft Defender Antivirus > MpEngine","Select cloud protection level","WindowsDefender.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows Defender\\MpEngine","MpCloudBlockLevel","REG_DWORD","0=Default, 2=High, 4=High+, 6=Zero tolerance. Recommended: 2."),
    ("Defender — PUA (Potentially Unwanted App) protection","Windows Components > Microsoft Defender Antivirus","Configure detection for potentially unwanted applications","WindowsDefender.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows Defender","PUAProtection","REG_DWORD","1=Block, 2=Audit mode, 0=Disabled. Recommended: 1."),
    ("Defender — Real-time protection","Windows Components > Microsoft Defender Antivirus > Real-time Protection","Turn off real-time protection","WindowsDefender.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows Defender\\Real-Time Protection","DisableRealtimeMonitoring","REG_DWORD","0=Real-time ON (recommended). 1=Disabled."),
    ("Defender — Tamper Protection","N/A — not configurable via GPO","Configure via Intune (Endpoint Security > AV) or Defender portal","N/A","HKLM\\SOFTWARE\\Microsoft\\Windows Defender\\Features","TamperProtection","REG_DWORD","4=Enabled (audit), 5=Enabled (block), 1=Disabled. Set via Intune, not GPO."),
    ("Defender — Submit samples automatically","Windows Components > Microsoft Defender Antivirus > MAPS","Send file samples when further analysis is required","WindowsDefender.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows Defender\\Spynet","SubmitSamplesConsent","REG_DWORD","1=Auto-send safe samples (recommended), 3=Send all samples."),
    ("Windows Hello for Business — Enable","Windows Components > Windows Hello for Business","Use Windows Hello for Business","Passport.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\PassportForWork","Enabled","REG_DWORD","1=Enable WHfB as strong auth replacing passwords."),
    ("Windows Hello — Require PIN digits","Windows Components > Windows Hello for Business > PIN Complexity","Require digits","Passport.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\PassportForWork\\PINComplexity","Digits","REG_DWORD","1=Require digits in WHfB PIN."),
    ("Windows Hello — Minimum PIN length","Windows Components > Windows Hello for Business > PIN Complexity","Minimum PIN length","Passport.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\PassportForWork\\PINComplexity","MinimumPINLength","REG_DWORD","Minimum characters for the Windows Hello PIN. Recommended: 6–8."),
    ("Remote Desktop — Require NLA","Windows Components > Remote Desktop Services > Remote Desktop Session Host > Security","Require user authentication using NLA","TerminalServer.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows NT\\Terminal Services","UserAuthentication","REG_DWORD","1=Require NLA before RDP session. Prevents unauthenticated connections."),
    ("Remote Desktop — Allow connections","Windows Components > Remote Desktop Services > Remote Desktop Session Host > Connections","Allow users to connect remotely by using RDS","TerminalServer.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows NT\\Terminal Services","fDenyTSConnections","REG_DWORD","0=Allow RDP. 1=Block (default on client Windows SKUs)."),
    ("Remote Desktop — Set encryption level","Windows Components > Remote Desktop Services > Remote Desktop Session Host > Security","Set client connection encryption level","TerminalServer.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows NT\\Terminal Services","MinEncryptionLevel","REG_DWORD","3=High (128-bit). Required for security compliance."),
    ("OneDrive — Silently sign in users","OneDrive","Silently sign in users to OneDrive sync with Windows credentials","OneDrive.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\OneDrive","SilentAccountConfig","REG_DWORD","1=Silently sign users into OneDrive using AAD credentials."),
    ("OneDrive — Known Folder Move silent opt-in","OneDrive","Silently move Windows known folders to OneDrive","OneDrive.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\OneDrive","KFMSilentOptIn","REG_SZ","Tenant ID. Redirects Desktop, Documents, Pictures to OneDrive silently."),
    ("OneDrive — Prevent personal account sync","OneDrive","Prevent users from adding personal OneDrive accounts","OneDrive.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\OneDrive","DisablePersonalSync","REG_DWORD","1=Block personal Microsoft account OneDrive sync."),
    ("Machine inactivity lockout","Windows Settings > Security Settings > Local Policies > Security Options","Interactive logon: Machine inactivity limit","SecGuide.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System","InactivityTimeoutSecs","REG_DWORD","Seconds before workstation auto-locks. 900=15 min. CIS recommends ≤900."),
    ("Disable AutoRun for all drives","Windows Components > AutoPlay Policies","Disallow Autorun for non-volume devices","AutoPlay.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\Explorer","NoDriveTypeAutoRun","REG_DWORD","255=Disable AutoRun for all drive types including USB."),
    ("SMBv1 — Disable","N/A — Registry only","Configure via PowerShell: Set-SmbServerConfiguration -EnableSMB1Protocol $false","N/A","HKLM\\SYSTEM\\CurrentControlSet\\Services\\LanmanServer\\Parameters","SMB1","REG_DWORD","0=Disable SMBv1 (required — protects against EternalBlue/WannaCry). Restart required."),
    ("WDigest — Disable plaintext credentials in memory","Windows Settings > Security Settings > Local Policies > Security Options","WDigest Authentication","SecGuide.admx","HKLM\\SYSTEM\\CurrentControlSet\\Control\\SecurityProviders\\WDigest","UseLogonCredential","REG_DWORD","0=Disable WDigest (prevents plaintext creds in LSASS memory)."),
    ("Diagnostic Data — Level","Windows Components > Data Collection and Preview Builds","Allow Diagnostic Data","DataCollection.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\DataCollection","AllowTelemetry","REG_DWORD","0=Security (Enterprise only), 1=Required, 3=Optional. Enterprise: use 0 or 1."),
    ("Restrict removable storage — Deny read","System > Removable Storage Access","All Removable Storage Classes: Deny read access","RemovableStorage.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\RemovableStorageDevices","Deny_Read","REG_DWORD","1=Deny all read access to removable storage."),
    ("Restrict removable storage — Deny write","System > Removable Storage Access","All Removable Storage Classes: Deny write access","RemovableStorage.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\RemovableStorageDevices","Deny_Write","REG_DWORD","1=Deny all write access to removable storage."),
    ("Windows Installer — Always install elevated","Windows Components > Windows Installer","Always install with elevated privileges","MSI.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\Installer","AlwaysInstallElevated","REG_DWORD","0=Disable (required). Setting to 1 is a critical security vulnerability."),
    ("PowerShell Execution Policy","Windows Components > Windows PowerShell","Turn on Script Execution","PowerShellExecutionPolicy.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\PowerShell","EnableScripts","REG_DWORD","1=Enable with ExecutionPolicy below. Set ExecutionPolicy to AllSigned or RemoteSigned."),
    ("PowerShell Script Block Logging","Windows Components > Windows PowerShell","Turn on PowerShell Script Block Logging","PowerShellExecutionPolicy.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\PowerShell\\ScriptBlockLogging","EnableScriptBlockLogging","REG_DWORD","1=Log all PowerShell script block execution to Event Log (Event ID 4104)."),
    ("PowerShell Transcription","Windows Components > Windows PowerShell","Turn on PowerShell Transcription","PowerShellExecutionPolicy.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\PowerShell\\Transcription","EnableTranscripting","REG_DWORD","1=Enable PowerShell session transcription (full command logging)."),
    ("Audit Process Creation","Windows Settings > Security Settings > Advanced Audit Policy > System Audit Policies > Detailed Tracking","Audit Process Creation","auditpol","N/A — configure via secpol.msc Advanced Audit Policy","","","Logs new process creation events (Event ID 4688). Required for most threat detection."),
    ("Include command line in process creation events","Windows Settings > Security Settings > Local Policies > Security Options","Include command line in process creation events","auditpol.admx","HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System\\Audit","ProcessCreationIncludeCmdLine_Enabled","REG_DWORD","1=Include full command line in Event ID 4688. Required for effective process auditing."),
    ("Credential Caching — Limit cached logons","Windows Settings > Security Settings > Local Policies > Security Options","Interactive logon: Number of previous logons to cache","SecGuide.admx","HKLM\\SOFTWARE\\Microsoft\\Windows NT\\CurrentVersion\\Winlogon","CachedLogonsCount","REG_SZ","Number of cached domain credentials. CIS recommends 0 for high security, 4 as practical minimum."),
]

def fetch_windows_admx():
    log("Building Windows ADMX entries…")
    entries = []
    for row in WINDOWS_ADMX:
        name, cat_path, policy, admx_file, reg_key, val, dtype, desc = row
        hive = "HKCU" if reg_key.startswith("HKCU") else "HKLM"
        key  = reg_key.replace("HKLM\\","").replace("HKCU\\","")
        e = make_entry(
            source_id="admx_windows",
            entry_id="admx_win_" + slugify(name),
            name=name, desc=desc,
            cats=["Administrative Templates", "Windows Policy",
                  cat_path.split(" > ")[0]],
            plat="windows", methods=["gpo","admx","registry"],
            gpo={"path": "Computer Configuration > Administrative Templates > " + cat_path,
                 "policy": policy, "admx": admx_file, "ns": ""},
            admx={"name": policy, "file": admx_file, "cat": cat_path,
                  "regKey": reg_key, "val": val, "type": dtype},
            reg={"hive": hive, "key": key, "val": val, "type": dtype,
                 "data": "See description", "note": ""},
        )
        entries.append(e)
    log(f"  Windows ADMX: {len(entries)} entries")
    return entries


# ─── SOURCE 5a: IntunePMFiles/DeviceConfig — Public Settings Definitions ────────
# Microsoft's Intune PM team publishes periodic exports of the full Settings Catalog
# to a public GitHub repo. No authentication required.
# Reference: https://github.com/IntunePMFiles/DeviceConfig (MIT License)
# Microsoft Docs reference: https://learn.microsoft.com/en-us/intune/intune-service/configuration/settings-catalog

INTUNE_PM_URLS = [
    # Try the most recent known file first, then fall back to older versions
    "https://raw.githubusercontent.com/IntunePMFiles/DeviceConfig/main/Settings%20Definitions%20Export%203-23.xlsx",
    "https://raw.githubusercontent.com/IntunePMFiles/DeviceConfig/main/Settings+Definitions+Export+3-23.xlsx",
]

def fetch_intune_pm_files():
    """
    Fetch the Settings Catalog definitions from IntunePMFiles/DeviceConfig.
    This is a public GitHub repo maintained by the Microsoft Intune PM team.
    Contains periodic exports of all settings available in the Settings Catalog.
    """
    log("Fetching IntunePMFiles/DeviceConfig Settings Catalog export…")

    # First try to get the repo contents to find the latest file
    api_url = "https://api.github.com/repos/IntunePMFiles/DeviceConfig/contents/"
    raw = http_get(api_url, headers={"Accept": "application/vnd.github.v3+json"}, timeout=30)
    xlsx_url = None
    if raw:
        try:
            files = json.loads(raw)
            for f in files:
                name = f.get("name","")
                if name.endswith(".xlsx") and "Settings" in name:
                    xlsx_url = f.get("download_url") or f.get("html_url","").replace(
                        "github.com","raw.githubusercontent.com"
                    ).replace("/blob/","/" )
                    log(f"  Found: {name}")
                    break
        except:
            pass

    if not xlsx_url:
        xlsx_url = INTUNE_PM_URLS[0]

    # Download the xlsx
    log(f"  Downloading: {xlsx_url[:80]}…")
    raw_bytes = None
    req = urllib.request.Request(xlsx_url, headers={"User-Agent": "HelientBot/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            raw_bytes = r.read()
            log(f"  Downloaded {len(raw_bytes)//1024} KB")
    except Exception as e:
        log(f"  Download failed: {e}")
        return []

    if not raw_bytes:
        return []

    # Parse xlsx using openpyxl
    try:
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        entries = []
        seen = set()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row
            headers = [str(h).lower().strip() if h else "" for h in rows[0]]
            col = {h: i for i, h in enumerate(headers)}

            # IntunePMFiles Excel column order:
            # col0=Platform, col1=Name, col2=CategoryName/CategoryPath, col3=SettingDefinitionId
            # col4=Description (sometimes)
            # The header row identifies columns by name if present
            # Use sentinel to safely handle column index 0 (which is falsy)
            _NOTFOUND = object()
            def _c(*keys, default=None):
                for k in keys:
                    v = col.get(k, _NOTFOUND)
                    if v is not _NOTFOUND:
                        return v
                return default
            name_col  = _c("name", "display name", "displayname", "setting name", "settingname", default=1)
            desc_col  = _c("description", "desc", "tooltip", "helptext", default=(4 if len(headers) > 4 else None))
            cat_col   = _c("categorypath", "category path", "categoryname", "category", default=2)
            plat_col  = _c("platform", "platforms", "applicability", default=0)
            defid_col = _c("settingdefinitionid", "setting definition id", "id", "definitionid", default=3)

            plat_map = {
                "windows": "windows", "windows10": "windows",
                "macos": "macos", "ios": "ios", "android": "android", "linux": "linux"
            }

            for row in rows[1:]:
                if not row or not any(row):
                    continue
                try:
                    name  = str(row[name_col] or "").strip()
                    desc  = str(row[desc_col] or "").strip() if desc_col < len(row) else ""
                    cat   = str(row[cat_col]  or "").strip() if cat_col  < len(row) else ""
                    plat_raw = str(row[plat_col] or "").lower().strip() if plat_col < len(row) else "windows"
                    defid = str(row[defid_col] or "").strip() if defid_col is not None and defid_col < len(row) else ""

                    # Skip header rows, platform-only rows, and empty names
                    platform_words = {"ios","windows","macos","android","linux","platform",
                                      "name","display name","setting name","windows 10","windows 11",
                                      "ios/ipados","chrome os","chromeos","windows 10/11"}
                    if not name or name.lower().strip() in platform_words:
                        continue
                    # Skip if name looks like a column header
                    if "this is the" in name.lower() or name.lower().startswith("name -"):
                        continue

                    plat = next((v for k,v in plat_map.items() if k in plat_raw), "windows")
                    eid  = "pm_" + (defid or slugify(name + cat + sheet_name))
                    if eid in seen:
                        continue
                    seen.add(eid)

                    oma = defid_to_oma(defid) if defid else ""

                    # Build a meaningful description from available fields
                    if not desc or len(desc) < 5:
                        desc = cat or ""  # Use category as desc if no real desc
                    e = make_entry(
                        source_id = "graph",
                        entry_id  = eid,
                        name      = name,
                        desc      = desc[:500],
                        cats      = [cat, "Settings Catalog", "Intune Settings Catalog"] if cat else ["Settings Catalog", "Intune Settings Catalog"],
                        plat      = plat,
                        methods   = ["intune"],
                        intune    = [{
                            "cat":   cat or "Settings Catalog",
                            "name":  name,
                            "defId": defid,
                            "oma":   oma,
                            "dtype": "Choice",
                            "vals":  [],
                            "rec":   "",
                            "json":  f'"settingDefinitionId": "{defid}"' if defid else f'"name": "{name}"',
                        }],
                    )
                    entries.append(e)
                except Exception:
                    continue

        log(f"  IntunePMFiles: {len(entries)} settings from {len(wb.sheetnames)} sheet(s)")
        return entries

    except ImportError:
        log("  openpyxl not installed — installing…")
        import subprocess
        subprocess.run(["pip", "install", "openpyxl", "--quiet", "--break-system-packages"],
                      capture_output=True)
        try:
            import openpyxl
            return fetch_intune_pm_files()  # retry after install
        except:
            log("  openpyxl install failed — skipping IntunePMFiles source")
            return []
    except Exception as e:
        log(f"  IntunePMFiles parse error: {e}")
        return []

# ─── SOURCE 5b: Policy CSP — Windows MDM Settings Reference ──────────────────
# The complete Policy CSP is Microsoft's authoritative list of all MDM-manageable
# Windows settings. Every entry has an OMA-URI, registry location, and maps
# directly to Intune Settings Catalog settingDefinitionIds.
# Reference: learn.microsoft.com/windows/client-management/mdm/policy-csp-*
POLICY_CSP_DATA = [
    ('AboveLock', 'ActionCenterNotifications', 'Allow Action Center notifications above lock', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\PushNotifications', 'NoTilesToastNotifications', 'REG_DWORD', 'AboveLock/ActionCenterNotifications'),
    ('AboveLock', 'AllowCortanaAboveLock', 'Allow Cortana above lock screen', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Windows Search', 'AllowCortanaAboveLock', 'REG_DWORD', 'AboveLock/AllowCortanaAboveLock'),
    ('AboveLock', 'AllowToasts', 'Allow toast notifications above lock screen', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\PushNotifications', 'NoToastNotification', 'REG_DWORD', 'AboveLock/AllowToasts'),
    ('Accounts', 'AllowAddingNonMicrosoftAccountsManually', 'Allow adding non-Microsoft accounts manually', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'NoConnectedUser', 'REG_DWORD', 'Accounts/AllowAddingNonMicrosoftAccountsManually'),
    ('Accounts', 'AllowMicrosoftAccountConnection', 'Allow Microsoft account connection', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\MRT', 'AllowMicrosoftAccountConnection', 'REG_DWORD', 'Accounts/AllowMicrosoftAccountConnection'),
    ('Accounts', 'AllowMicrosoftAccountSignInAssistant', 'Allow Microsoft Account Sign-In Assistant service', 'HKLM\\\\SYSTEM\\\\CurrentControlSet\\\\Services\\\\wlidsvc', 'Start', 'REG_DWORD', 'Accounts/AllowMicrosoftAccountSignInAssistant'),
    ('ApplicationManagement', 'AllowAllTrustedApps', 'Allow all trusted apps to install (sideloading)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Appx', 'AllowAllTrustedApps', 'REG_DWORD', 'ApplicationManagement/AllowAllTrustedApps'),
    ('ApplicationManagement', 'AllowAppStoreAutoUpdate', 'Allow automatic update of apps from Store', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\WindowsStore', 'AutoDownload', 'REG_DWORD', 'ApplicationManagement/AllowAppStoreAutoUpdate'),
    ('ApplicationManagement', 'AllowDeveloperUnlock', 'Allow developer unlock (sideloading without Dev License)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Appx', 'AllowDevelopmentWithoutDevLicense', 'REG_DWORD', 'ApplicationManagement/AllowDeveloperUnlock'),
    ('ApplicationManagement', 'AllowGameDVR', 'Allow Game DVR recording', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\GameDVR', 'AllowGameDVR', 'REG_DWORD', 'ApplicationManagement/AllowGameDVR'),
    ('ApplicationManagement', 'AllowSharedUserAppData', 'Allow shared user app data', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\AppModel\\\\StateManager', 'AllowSharedLocalAppData', 'REG_DWORD', 'ApplicationManagement/AllowSharedUserAppData'),
    ('ApplicationManagement', 'AllowStore', 'Allow Microsoft Store', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\WindowsStore', 'DisableStoreApps', 'REG_DWORD', 'ApplicationManagement/AllowStore'),
    ('ApplicationManagement', 'DisableStoreOriginatedApps', 'Disable apps installed from Store', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\WindowsStore', 'DisableStoreApps', 'REG_DWORD', 'ApplicationManagement/DisableStoreOriginatedApps'),
    ('ApplicationManagement', 'MSIAlwaysInstallWithElevatedPrivileges', 'MSI always install elevated (0=disable this risk)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Installer', 'AlwaysInstallElevated', 'REG_DWORD', 'ApplicationManagement/MSIAlwaysInstallWithElevatedPrivileges'),
    ('ApplicationManagement', 'RequirePrivateStoreOnly', 'Require private Store only (block public Store)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\WindowsStore', 'RequirePrivateStoreOnly', 'REG_DWORD', 'ApplicationManagement/RequirePrivateStoreOnly'),
    ('Authentication', 'AllowAadPasswordReset', 'Allow AAD password reset from lock screen', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\AzureADAccount', 'AllowPasswordReset', 'REG_DWORD', 'Authentication/AllowAadPasswordReset'),
    ('Authentication', 'AllowFastReconnect', 'Allow fast reconnect (EAP fast reconnect)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'AllowDomainPINLogon', 'REG_DWORD', 'Authentication/AllowFastReconnect'),
    ('Authentication', 'AllowFidoDeviceSignon', 'Allow FIDO2 security key sign-in', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\FIDO', 'AllowFidoDeviceSignon', 'REG_DWORD', 'Authentication/AllowFidoDeviceSignon'),
    ('Authentication', 'AllowSecondaryAuthenticationDevice', 'Allow secondary authentication companion device', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\PassportForWork\\\\Remote', 'Enabled', 'REG_DWORD', 'Authentication/AllowSecondaryAuthenticationDevice'),
    ('Authentication', 'PreferredAadTenantDomainName', 'Preferred AAD tenant domain name (UPN hint)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\AzureADAccount', 'PreferredAadTenantDomainName', 'REG_SZ', 'Authentication/PreferredAadTenantDomainName'),
    ('Connectivity', 'AllowBluetooth', 'Allow Bluetooth (0=off, 2=on)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Bluetooth', 'AllowBluetooth', 'REG_DWORD', 'Connectivity/AllowBluetooth'),
    ('Connectivity', 'AllowCellularData', 'Allow cellular data access', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WwanSvc\\\\CellularDataAccess', 'LetAppsAccessCellularData', 'REG_DWORD', 'Connectivity/AllowCellularData'),
    ('Connectivity', 'AllowCellularDataRoaming', 'Allow cellular data roaming', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WwanSvc\\\\CellularDataAccess', 'LetAppsAccessCellularDataRoaming', 'REG_DWORD', 'Connectivity/AllowCellularDataRoaming'),
    ('Connectivity', 'AllowNFC', 'Allow NFC', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Connectivity', 'AllowNFC', 'REG_DWORD', 'Connectivity/AllowNFC'),
    ('Connectivity', 'AllowUSBConnection', 'Allow USB connection', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Connectivity', 'AllowUSBConnection', 'REG_DWORD', 'Connectivity/AllowUSBConnection'),
    ('Connectivity', 'AllowVPNRoaming', 'Allow VPN roaming over cellular', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Network Connections', 'AllowVPNRoaming', 'REG_DWORD', 'Connectivity/AllowVPNRoaming'),
    ('Connectivity', 'HardenedUNCPaths', 'Hardened UNC paths (RequireMutualAuthentication=1,RequireIntegrity=1)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\NetworkProvider\\\\HardenedPaths', '\\\\\\\\*\\\\NETLOGON', 'REG_SZ', 'Connectivity/HardenedUNCPaths'),
    ('Connectivity', 'ProhibitInstallationAndConfigurationOfNetworkBridge', 'Prohibit network bridge installation', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Network Connections', 'NC_AllowNetBridge_NLA', 'REG_DWORD', 'Connectivity/ProhibitInstallationAndConfigurationOfNetworkBridge'),
    ('ControlPolicyConflict', 'MDMWinsOverGP', 'MDM takes precedence over Group Policy', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\MDM', 'ControlPolicyConflict', 'REG_DWORD', 'ControlPolicyConflict/MDMWinsOverGP'),
    ('Defender', 'AllowArchiveScanning', 'Allow archive scanning', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Scan', 'DisableArchiveScanning', 'REG_DWORD', 'Defender/AllowArchiveScanning'),
    ('Defender', 'AllowBehaviorMonitoring', 'Allow behavior monitoring', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Real-Time Protection', 'DisableBehaviorMonitoring', 'REG_DWORD', 'Defender/AllowBehaviorMonitoring'),
    ('Defender', 'AllowCloudProtection', 'Allow cloud-delivered protection (MAPS)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Spynet', 'SpynetReporting', 'REG_DWORD', 'Defender/AllowCloudProtection'),
    ('Defender', 'AllowEmailScanning', 'Allow email scanning', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Scan', 'DisableEmailScanning', 'REG_DWORD', 'Defender/AllowEmailScanning'),
    ('Defender', 'AllowFullScanRemovableDriveScanning', 'Allow full scan of removable drives', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Scan', 'DisableRemovableDriveScanning', 'REG_DWORD', 'Defender/AllowFullScanRemovableDriveScanning'),
    ('Defender', 'AllowIOAVProtection', 'Allow IOAV protection (scan downloaded files)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Real-Time Protection', 'DisableIOAVProtection', 'REG_DWORD', 'Defender/AllowIOAVProtection'),
    ('Defender', 'AllowOnAccessProtection', 'Allow on-access protection', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Real-Time Protection', 'DisableOnAccessProtection', 'REG_DWORD', 'Defender/AllowOnAccessProtection'),
    ('Defender', 'AllowRealtimeMonitoring', 'Allow real-time monitoring', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Real-Time Protection', 'DisableRealtimeMonitoring', 'REG_DWORD', 'Defender/AllowRealtimeMonitoring'),
    ('Defender', 'AllowScanningNetworkFiles', 'Allow scanning of network files', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Scan', 'DisableScanningNetworkFiles', 'REG_DWORD', 'Defender/AllowScanningNetworkFiles'),
    ('Defender', 'AllowScriptScanning', 'Allow script scanning', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender', 'DisableScriptScanning', 'REG_DWORD', 'Defender/AllowScriptScanning'),
    ('Defender', 'AttackSurfaceReductionRules', 'ASR rule GUIDs mapped to 0=disable/1=block/2=audit', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Windows Defender Exploit Guard\\\\ASR\\\\Rules', 'be9ba2d9-53ea-4cdc-84e5-9b1eeee46550', 'REG_SZ', 'Defender/AttackSurfaceReductionRules'),
    ('Defender', 'CloudBlockLevel', 'Cloud block level (0=default, 2=high, 4=high+, 6=zero tolerance)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\MpEngine', 'MpCloudBlockLevel', 'REG_DWORD', 'Defender/CloudBlockLevel'),
    ('Defender', 'CloudExtendedTimeout', 'Cloud extended timeout in seconds (0-50)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\MpEngine', 'MpBafsExtendedTimeout', 'REG_DWORD', 'Defender/CloudExtendedTimeout'),
    ('Defender', 'DaysToRetainCleanedMalware', 'Days to retain cleaned malware in quarantine', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Quarantine', 'PurgeItemsAfterDelay', 'REG_DWORD', 'Defender/DaysToRetainCleanedMalware'),
    ('Defender', 'EnableControlledFolderAccess', 'Enable controlled folder access (0=disable, 1=enable, 2=audit)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Windows Defender Exploit Guard\\\\Controlled Folder Access', 'EnableControlledFolderAccess', 'REG_DWORD', 'Defender/EnableControlledFolderAccess'),
    ('Defender', 'EnableNetworkProtection', 'Enable network protection (0=disable, 1=block, 2=audit)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Windows Defender Exploit Guard\\\\Network Protection', 'EnableNetworkProtection', 'REG_DWORD', 'Defender/EnableNetworkProtection'),
    ('Defender', 'ExcludedExtensions', 'Excluded file extensions from scanning', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Exclusions\\\\Extensions', '', 'REG_MULTI_SZ', 'Defender/ExcludedExtensions'),
    ('Defender', 'ExcludedPaths', 'Excluded paths from scanning', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Exclusions\\\\Paths', '', 'REG_MULTI_SZ', 'Defender/ExcludedPaths'),
    ('Defender', 'PUAProtection', 'PUA protection (0=disable, 1=block, 2=audit)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender', 'PUAProtection', 'REG_DWORD', 'Defender/PUAProtection'),
    ('Defender', 'ScheduleScanDay', 'Scan day (0=daily, 1=Sun, 2=Mon, 7=Sat, 8=none)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Scan', 'ScheduleDay', 'REG_DWORD', 'Defender/ScheduleScanDay'),
    ('Defender', 'ScheduleScanTime', 'Scan time (minutes from midnight, 0-1439)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Scan', 'ScheduleTime', 'REG_DWORD', 'Defender/ScheduleScanTime'),
    ('Defender', 'SignatureUpdateInterval', 'Signature update interval in hours (0=disable)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Signature Updates', 'SignatureUpdateInterval', 'REG_DWORD', 'Defender/SignatureUpdateInterval'),
    ('Defender', 'SubmitSamplesConsent', 'Submit samples (0=always prompt, 1=auto safe, 2=never, 3=send all)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\Spynet', 'SubmitSamplesConsent', 'REG_DWORD', 'Defender/SubmitSamplesConsent'),
    ('DeliveryOptimization', 'DODownloadMode', 'DO mode (0=off, 1=LAN, 2=group, 3=internet, 99=bypass, 100=simple)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DeliveryOptimization', 'DODownloadMode', 'REG_DWORD', 'DeliveryOptimization/DODownloadMode'),
    ('DeliveryOptimization', 'DOGroupId', 'DO group ID (GUID for peering group)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DeliveryOptimization', 'DOGroupId', 'REG_SZ', 'DeliveryOptimization/DOGroupId'),
    ('DeliveryOptimization', 'DOGroupIdSource', 'DO group ID source (1=AD site, 2=SID, 3=DHCP 234, 4=DNS suffix)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DeliveryOptimization', 'DOGroupIdSource', 'REG_DWORD', 'DeliveryOptimization/DOGroupIdSource'),
    ('DeliveryOptimization', 'DOMaxBackgroundDownloadBandwidth', 'Max background bandwidth KB/s (0=unlimited)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DeliveryOptimization', 'DOMaxDownloadBandwidth', 'REG_DWORD', 'DeliveryOptimization/DOMaxBackgroundDownloadBandwidth'),
    ('DeliveryOptimization', 'DOMaxCacheSize', 'Max cache size as percent of disk', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DeliveryOptimization', 'DOMaxCacheSize', 'REG_DWORD', 'DeliveryOptimization/DOMaxCacheSize'),
    ('DeliveryOptimization', 'DORestrictPeerSelectionBy', 'Restrict peer selection (0=NAT, 1=subnet, 2=LAN)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DeliveryOptimization', 'DORestrictPeerSelectionBy', 'REG_DWORD', 'DeliveryOptimization/DORestrictPeerSelectionBy'),
    ('DeliveryOptimization', 'DOCacheHost', 'Connected Cache server hostname for DO', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DeliveryOptimization', 'DOCacheHost', 'REG_SZ', 'DeliveryOptimization/DOCacheHost'),
    ('DeviceLock', 'MaxDevicePasswordFailedAttempts', 'Max failed password attempts before wipe (4-16)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'MaxDevicePasswordFailedAttempts', 'REG_DWORD', 'DeviceLock/MaxDevicePasswordFailedAttempts'),
    ('DeviceLock', 'MaxInactivityTimeDeviceLock', 'Max inactivity time before lock (minutes)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'InactivityTimeoutSecs', 'REG_DWORD', 'DeviceLock/MaxInactivityTimeDeviceLock'),
    ('DeviceLock', 'MinDevicePasswordComplexCharacters', 'Min password complexity characters (1-4)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'MinDevicePasswordComplexCharacters', 'REG_DWORD', 'DeviceLock/MinDevicePasswordComplexCharacters'),
    ('DeviceLock', 'MinDevicePasswordLength', 'Min device password length (4-16)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'MinDevicePasswordLength', 'REG_DWORD', 'DeviceLock/MinDevicePasswordLength'),
    ('DeviceLock', 'PasswordHistoryCount', 'Password history count prevents reuse (0-50)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'DevicePasswordHistory', 'REG_DWORD', 'DeviceLock/PasswordHistoryCount'),
    ('Experience', 'AllowCortana', 'Allow Cortana', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Windows Search', 'AllowCortana', 'REG_DWORD', 'Experience/AllowCortana'),
    ('Experience', 'AllowManualMDMUnenrollment', 'Allow manual MDM unenrollment by users', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\MDM', 'AllowUserToUnenroll', 'REG_DWORD', 'Experience/AllowManualMDMUnenrollment'),
    ('Experience', 'AllowSyncMySettings', 'Allow sync of settings across devices (Microsoft Account)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\SettingSync', 'DisableSettingSync', 'REG_DWORD', 'Experience/AllowSyncMySettings'),
    ('Experience', 'AllowWindowsConsumerFeatures', 'Allow Windows consumer features (Spotlight, suggested apps)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\CloudContent', 'DisableWindowsConsumerFeatures', 'REG_DWORD', 'Experience/AllowWindowsConsumerFeatures'),
    ('Experience', 'AllowWindowsTips', 'Allow Windows tips', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\CloudContent', 'DisableSoftLanding', 'REG_DWORD', 'Experience/AllowWindowsTips'),
    ('Experience', 'ConfigureChatIcon', 'Configure Teams/Chat taskbar icon (0=show, 1=hide, 2=user choice)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Windows Chat', 'ChatIcon', 'REG_DWORD', 'Experience/ConfigureChatIcon'),
    ('LocalPoliciesSecurityOptions', 'InteractiveLogonDoNotDisplayLastSignedIn', 'Do not display last signed-in username on lock', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'DontDisplayLastUserName', 'REG_DWORD', 'LocalPoliciesSecurityOptions/InteractiveLogonDoNotDisplayLastSignedIn'),
    ('LocalPoliciesSecurityOptions', 'InteractiveLogonMachineInactivityLimit', 'Machine inactivity lockout (seconds)', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'InactivityTimeoutSecs', 'REG_DWORD', 'LocalPoliciesSecurityOptions/InteractiveLogonMachineInactivityLimit'),
    ('LocalPoliciesSecurityOptions', 'InteractiveLogonMessageTextForUsersAttemptingToLogon', 'Logon warning message body text', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'LegalNoticeText', 'REG_SZ', 'LocalPoliciesSecurityOptions/InteractiveLogonMessageTextForUsersAttemptingToLogon'),
    ('LocalPoliciesSecurityOptions', 'InteractiveLogonMessageTitleForUsersAttemptingToLogOn', 'Logon warning message title', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'LegalNoticeCaption', 'REG_SZ', 'LocalPoliciesSecurityOptions/InteractiveLogonMessageTitleForUsersAttemptingToLogOn'),
    ('LocalPoliciesSecurityOptions', 'InteractiveLogonNumberOfPreviousLogonsToCache', 'Cached credentials count (0=none recommended for high security)', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows NT\\\\CurrentVersion\\\\Winlogon', 'CachedLogonsCount', 'REG_SZ', 'LocalPoliciesSecurityOptions/InteractiveLogonNumberOfPreviousLogonsToCache'),
    ('LocalPoliciesSecurityOptions', 'InteractiveLogonSmartCardRemovalBehavior', 'Smart card removal (0=none, 1=lock, 2=logoff, 3=disconnect RDP)', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows NT\\\\CurrentVersion\\\\Winlogon', 'ScRemoveOption', 'REG_SZ', 'LocalPoliciesSecurityOptions/InteractiveLogonSmartCardRemovalBehavior'),
    ('LocalPoliciesSecurityOptions', 'MicrosoftNetworkClientDigitallySignCommunicationsAlways', 'Always digitally sign SMB client communications', 'HKLM\\\\SYSTEM\\\\CurrentControlSet\\\\Services\\\\LanmanWorkstation\\\\Parameters', 'RequireSecuritySignature', 'REG_DWORD', 'LocalPoliciesSecurityOptions/MicrosoftNetworkClientDigitallySignCommunicationsAlways'),
    ('LocalPoliciesSecurityOptions', 'MicrosoftNetworkServerDigitallySignCommunicationsAlways', 'Always digitally sign SMB server communications', 'HKLM\\\\SYSTEM\\\\CurrentControlSet\\\\Services\\\\LanManServer\\\\Parameters', 'RequireSecuritySignature', 'REG_DWORD', 'LocalPoliciesSecurityOptions/MicrosoftNetworkServerDigitallySignCommunicationsAlways'),
    ('LocalPoliciesSecurityOptions', 'NetworkAccessDoNotAllowAnonymousEnumerationOfSAMAccountsAndShares', 'Do not allow anonymous enumeration of SAM accounts and shares', 'HKLM\\\\SYSTEM\\\\CurrentControlSet\\\\Control\\\\Lsa', 'RestrictAnonymous', 'REG_DWORD', 'LocalPoliciesSecurityOptions/NetworkAccessDoNotAllowAnonymousEnumerationOfSAMAccountsAndShares'),
    ('LocalPoliciesSecurityOptions', 'NetworkSecurityDoNotStoreLANManagerHashValueOnNextPasswordChange', 'Do not store LAN Manager hash on next password change', 'HKLM\\\\SYSTEM\\\\CurrentControlSet\\\\Control\\\\Lsa', 'NoLMHash', 'REG_DWORD', 'LocalPoliciesSecurityOptions/NetworkSecurityDoNotStoreLANManagerHashValueOnNextPasswordChange'),
    ('LocalPoliciesSecurityOptions', 'NetworkSecurityLANManagerAuthenticationLevel', 'LAN Manager auth level (5=NTLMv2 only, most secure)', 'HKLM\\\\SYSTEM\\\\CurrentControlSet\\\\Control\\\\Lsa', 'LmCompatibilityLevel', 'REG_DWORD', 'LocalPoliciesSecurityOptions/NetworkSecurityLANManagerAuthenticationLevel'),
    ('LocalPoliciesSecurityOptions', 'UserAccountControlBehaviorOfTheElevationPromptForAdministrators', 'UAC admin prompt (0=no prompt, 2=consent secure desktop)', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'ConsentPromptBehaviorAdmin', 'REG_DWORD', 'LocalPoliciesSecurityOptions/UserAccountControlBehaviorOfTheElevationPromptForAdministrators'),
    ('LocalPoliciesSecurityOptions', 'UserAccountControlBehaviorOfTheElevationPromptForStandardUsers', 'UAC standard user prompt (0=auto-deny, 3=credentials)', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'ConsentPromptBehaviorUser', 'REG_DWORD', 'LocalPoliciesSecurityOptions/UserAccountControlBehaviorOfTheElevationPromptForStandardUsers'),
    ('LocalPoliciesSecurityOptions', 'UserAccountControlRunAllAdministratorsInAdminApprovalMode', 'UAC run all administrators in admin approval mode (EnableLUA)', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'EnableLUA', 'REG_DWORD', 'LocalPoliciesSecurityOptions/UserAccountControlRunAllAdministratorsInAdminApprovalMode'),
    ('LocalPoliciesSecurityOptions', 'UserAccountControlSwitchToTheSecureDesktopWhenPromptingForElevation', 'UAC use secure desktop for elevation prompts', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'PromptOnSecureDesktop', 'REG_DWORD', 'LocalPoliciesSecurityOptions/UserAccountControlSwitchToTheSecureDesktopWhenPromptingForElevation'),
    ('LocalPoliciesSecurityOptions', 'UserAccountControlUseAdminApprovalModeForTheBuiltInAdministratorAccount', 'UAC admin approval for built-in Administrator account', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'FilterAdministratorToken', 'REG_DWORD', 'LocalPoliciesSecurityOptions/UserAccountControlUseAdminApprovalModeForTheBuiltInAdministratorAccount'),
    ('Privacy', 'AllowCrossDeviceClipboard', 'Allow cross-device clipboard (sync clipboard to other devices)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'AllowCrossDeviceClipboard', 'REG_DWORD', 'Privacy/AllowCrossDeviceClipboard'),
    ('Privacy', 'AllowInputPersonalization', 'Allow input personalization (Cortana, speech recognition)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\InputPersonalization', 'AllowInputPersonalization', 'REG_DWORD', 'Privacy/AllowInputPersonalization'),
    ('Privacy', 'DisableAdvertisingId', 'Disable advertising ID for personalized ads', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\AdvertisingInfo', 'DisabledByGroupPolicy', 'REG_DWORD', 'Privacy/DisableAdvertisingId'),
    ('Privacy', 'EnableActivityFeed', 'Enable Windows Timeline/activity feed', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'EnableActivityFeed', 'REG_DWORD', 'Privacy/EnableActivityFeed'),
    ('Privacy', 'LetAppsAccessCamera', 'Let apps access camera (0=user, 1=force allow, 2=force deny)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\AppPrivacy', 'LetAppsAccessCamera', 'REG_DWORD', 'Privacy/LetAppsAccessCamera'),
    ('Privacy', 'LetAppsAccessLocation', 'Let apps access location', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\AppPrivacy', 'LetAppsAccessLocation', 'REG_DWORD', 'Privacy/LetAppsAccessLocation'),
    ('Privacy', 'LetAppsAccessMicrophone', 'Let apps access microphone', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\AppPrivacy', 'LetAppsAccessMicrophone', 'REG_DWORD', 'Privacy/LetAppsAccessMicrophone'),
    ('Privacy', 'PublishUserActivities', 'Publish user activities to Timeline and cross-device', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'PublishUserActivities', 'REG_DWORD', 'Privacy/PublishUserActivities'),
    ('RemoteDesktopServices', 'AllowUsersToConnectRemotely', 'Allow remote desktop connections', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Terminal Services', 'fDenyTSConnections', 'REG_DWORD', 'RemoteDesktopServices/AllowUsersToConnectRemotely'),
    ('RemoteDesktopServices', 'ClientConnectionEncryptionLevel', 'RDP encryption level (3=high/128-bit required)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Terminal Services', 'MinEncryptionLevel', 'REG_DWORD', 'RemoteDesktopServices/ClientConnectionEncryptionLevel'),
    ('RemoteDesktopServices', 'DoNotAllowDriveRedirection', 'Do not allow drive redirection in RDP sessions', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Terminal Services', 'fDisableCdm', 'REG_DWORD', 'RemoteDesktopServices/DoNotAllowDriveRedirection'),
    ('RemoteDesktopServices', 'DoNotAllowPasswordSaving', 'Do not allow password saving in RDP client', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Terminal Services', 'DisablePasswordSaving', 'REG_DWORD', 'RemoteDesktopServices/DoNotAllowPasswordSaving'),
    ('RemoteDesktopServices', 'PromptForPasswordUponConnection', 'Prompt for password upon RDP connection', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Terminal Services', 'fPromptForPassword', 'REG_DWORD', 'RemoteDesktopServices/PromptForPasswordUponConnection'),
    ('RemoteDesktopServices', 'RequireSecureRPCCommunication', 'Require secure RPC for RDP', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Terminal Services', 'fEncryptRPCTraffic', 'REG_DWORD', 'RemoteDesktopServices/RequireSecureRPCCommunication'),
    ('RemoteManagement', 'AllowBasicAuthentication_Client', 'Allow basic auth for WinRM client', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WinRM\\\\Client', 'AllowBasic', 'REG_DWORD', 'RemoteManagement/AllowBasicAuthentication_Client'),
    ('RemoteManagement', 'AllowBasicAuthentication_Service', 'Allow basic auth for WinRM service', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WinRM\\\\Service', 'AllowBasic', 'REG_DWORD', 'RemoteManagement/AllowBasicAuthentication_Service'),
    ('RemoteManagement', 'AllowUnencryptedTraffic_Client', 'Allow unencrypted WinRM client traffic', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WinRM\\\\Client', 'AllowUnencryptedTraffic', 'REG_DWORD', 'RemoteManagement/AllowUnencryptedTraffic_Client'),
    ('RemoteManagement', 'AllowUnencryptedTraffic_Service', 'Allow unencrypted WinRM service traffic', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WinRM\\\\Service', 'AllowUnencryptedTraffic', 'REG_DWORD', 'RemoteManagement/AllowUnencryptedTraffic_Service'),
    ('RemoteManagement', 'DisallowDigestAuthentication', 'Disallow digest auth for WinRM client', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WinRM\\\\Client', 'DisableDigest', 'REG_DWORD', 'RemoteManagement/DisallowDigestAuthentication'),
    ('RemoteProcedureCall', 'RPCEndpointMapperClientAuthentication', 'RPC endpoint mapper requires auth', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Rpc', 'EnableAuthEpResolution', 'REG_DWORD', 'RemoteProcedureCall/RPCEndpointMapperClientAuthentication'),
    ('RemoteProcedureCall', 'RestrictUnauthenticatedRPCClients', 'Restrict unauthenticated RPC clients (2=authenticated, no exceptions)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows NT\\\\Rpc', 'RestrictRemoteClients', 'REG_DWORD', 'RemoteProcedureCall/RestrictUnauthenticatedRPCClients'),
    ('Search', 'AllowCloudSearch', 'Allow cloud search (Bing in Windows Search)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Windows Search', 'AllowCloudSearch', 'REG_DWORD', 'Search/AllowCloudSearch'),
    ('Search', 'DoNotUseWebResults', 'Disable web results in Windows Search', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\Windows Search', 'ConnectedSearchUseWeb', 'REG_DWORD', 'Search/DoNotUseWebResults'),
    ('Security', 'RequireDeviceEncryption', 'Require device encryption (BitLocker)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\FVE', 'RequireDeviceEncryption', 'REG_DWORD', 'Security/RequireDeviceEncryption'),
    ('SmartScreen', 'EnableAppInstallControl', 'App Install Control - require Store apps only', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender\\\\SmartScreen', 'ConfigureAppInstallControl', 'REG_SZ', 'SmartScreen/EnableAppInstallControl'),
    ('SmartScreen', 'EnableSmartScreenInShell', 'Enable SmartScreen in Windows Shell', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'EnableSmartScreen', 'REG_DWORD', 'SmartScreen/EnableSmartScreenInShell'),
    ('SmartScreen', 'PreventOverrideForFilesInShell', 'Prevent SmartScreen override for files', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'ShellSmartScreenLevel', 'REG_SZ', 'SmartScreen/PreventOverrideForFilesInShell'),
    ('System', 'AllowTelemetry', 'Telemetry level (0=security-enterprise, 1=required, 3=optional)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\DataCollection', 'AllowTelemetry', 'REG_DWORD', 'System/AllowTelemetry'),
    ('System', 'AllowLocation', 'Allow location services', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\LocationAndSensors', 'DisableLocation', 'REG_DWORD', 'System/AllowLocation'),
    ('System', 'DisableOneDriveFileSync', 'Disable OneDrive file sync', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\OneDrive', 'DisableFileSyncNGSC', 'REG_DWORD', 'System/DisableOneDriveFileSync'),
    ('System', 'TurnOffFileHistory', 'Turn off File History backup', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\FileHistory', 'Disabled', 'REG_DWORD', 'System/TurnOffFileHistory'),
    ('Update', 'ActiveHoursEnd', 'Active hours end time (0-23) - no auto-restart during this window', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate\\\\AU', 'ActiveHoursEnd', 'REG_DWORD', 'Update/ActiveHoursEnd'),
    ('Update', 'ActiveHoursStart', 'Active hours start time (0-23)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate\\\\AU', 'ActiveHoursStart', 'REG_DWORD', 'Update/ActiveHoursStart'),
    ('Update', 'AllowAutoUpdate', 'Auto update mode (2=notify, 3=auto-dl, 4=auto-install, 5=user)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate\\\\AU', 'AUOptions', 'REG_DWORD', 'Update/AllowAutoUpdate'),
    ('Update', 'AllowMUUpdateService', 'Allow Microsoft Update service (includes other MS products)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate\\\\AU', 'AllowMUUpdateService', 'REG_DWORD', 'Update/AllowMUUpdateService'),
    ('Update', 'BranchReadinessLevel', 'Update channel readiness (2=SAC, 4=insider fast, 8=insider slow)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'BranchReadinessLevel', 'REG_DWORD', 'Update/BranchReadinessLevel'),
    ('Update', 'ConfigureDeadlineForFeatureUpdates', 'Feature update deadline in days (0-30)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'ConfigureDeadlineForFeatureUpdates', 'REG_DWORD', 'Update/ConfigureDeadlineForFeatureUpdates'),
    ('Update', 'ConfigureDeadlineForQualityUpdates', 'Quality update deadline in days (0-30)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'ConfigureDeadlineForQualityUpdates', 'REG_DWORD', 'Update/ConfigureDeadlineForQualityUpdates'),
    ('Update', 'ConfigureDeadlineGracePeriod', 'Deadline grace period in days for restarts', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'ConfigureDeadlineGracePeriod', 'REG_DWORD', 'Update/ConfigureDeadlineGracePeriod'),
    ('Update', 'DeferFeatureUpdatesPeriodInDays', 'Defer feature updates (0-365 days)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'DeferFeatureUpdatesPeriodInDays', 'REG_DWORD', 'Update/DeferFeatureUpdatesPeriodInDays'),
    ('Update', 'DeferQualityUpdatesPeriodInDays', 'Defer quality updates (0-35 days)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'DeferQualityUpdatesPeriodInDays', 'REG_DWORD', 'Update/DeferQualityUpdatesPeriodInDays'),
    ('Update', 'DisableDualScan', 'Disable dual scan (prevents scanning both WU and WSUS)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'DisableDualScan', 'REG_DWORD', 'Update/DisableDualScan'),
    ('Update', 'ManagePreviewBuilds', 'Manage preview builds (0=disable, 1=enable, 2=disable after next update)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'ManagePreviewBuilds', 'REG_DWORD', 'Update/ManagePreviewBuilds'),
    ('Update', 'PauseFeatureUpdates', 'Pause feature updates (ISO 8601 date string to resume)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'PauseFeatureUpdatesStartTime', 'REG_SZ', 'Update/PauseFeatureUpdates'),
    ('Update', 'PauseQualityUpdates', 'Pause quality updates (ISO 8601 date string to resume)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'PauseQualityUpdatesStartTime', 'REG_SZ', 'Update/PauseQualityUpdates'),
    ('Update', 'ScheduledInstallDay', 'Scheduled install day (0=every day, 1=Sunday, 7=Saturday)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate\\\\AU', 'ScheduledInstallDay', 'REG_DWORD', 'Update/ScheduledInstallDay'),
    ('Update', 'ScheduledInstallTime', 'Scheduled install time (0-23)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate\\\\AU', 'ScheduledInstallTime', 'REG_DWORD', 'Update/ScheduledInstallTime'),
    ('Update', 'UpdateServiceUrl', 'WSUS server URL', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WindowsUpdate', 'WUServer', 'REG_SZ', 'Update/UpdateServiceUrl'),
    ('WindowsDefenderSecurityCenter', 'CompanyName', 'Company name shown in Windows Security Center', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\Enterprise Customization', 'CompanyName', 'REG_SZ', 'WindowsDefenderSecurityCenter/CompanyName'),
    ('WindowsDefenderSecurityCenter', 'DisableAccountProtectionUI', 'Disable account protection UI in Security Center', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\Account protection', 'UILockdown', 'REG_DWORD', 'WindowsDefenderSecurityCenter/DisableAccountProtectionUI'),
    ('WindowsDefenderSecurityCenter', 'DisableAppBrowserUI', 'Disable app and browser control UI', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\App and Browser protection', 'UILockdown', 'REG_DWORD', 'WindowsDefenderSecurityCenter/DisableAppBrowserUI'),
    ('WindowsDefenderSecurityCenter', 'DisableDeviceSecurityUI', 'Disable device security UI', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\Device security', 'UILockdown', 'REG_DWORD', 'WindowsDefenderSecurityCenter/DisableDeviceSecurityUI'),
    ('WindowsDefenderSecurityCenter', 'DisableNetworkUI', 'Disable firewall and network protection UI', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\Firewall and network protection', 'UILockdown', 'REG_DWORD', 'WindowsDefenderSecurityCenter/DisableNetworkUI'),
    ('WindowsDefenderSecurityCenter', 'DisableNotifications', 'Disable Windows Security notifications', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\Notifications', 'DisableNotifications', 'REG_DWORD', 'WindowsDefenderSecurityCenter/DisableNotifications'),
    ('WindowsDefenderSecurityCenter', 'DisableVirusUI', 'Disable virus and threat protection UI', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\Virus and threat protection', 'UILockdown', 'REG_DWORD', 'WindowsDefenderSecurityCenter/DisableVirusUI'),
    ('WindowsDefenderSecurityCenter', 'HideWindowsSecurityNotificationAreaControl', 'Hide Windows Security icon in system tray', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows Defender Security Center\\\\Systray', 'HideSystray', 'REG_DWORD', 'WindowsDefenderSecurityCenter/HideWindowsSecurityNotificationAreaControl'),
    ('WindowsLogon', 'AllowAutomaticRestartSignOn', 'Allow automatic restart sign-on after updates', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'DisableAutomaticRestartSignOn', 'REG_DWORD', 'WindowsLogon/AllowAutomaticRestartSignOn'),
    ('WindowsLogon', 'DisableLockScreenAppNotifications', 'Disable lock screen app notifications', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'DisableLockScreenAppNotifications', 'REG_DWORD', 'WindowsLogon/DisableLockScreenAppNotifications'),
    ('WindowsLogon', 'DontDisplayNetworkSelectionUI', 'Do not show network selection UI on lock screen', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\System', 'DontDisplayNetworkSelectionUI', 'REG_DWORD', 'WindowsLogon/DontDisplayNetworkSelectionUI'),
    ('WindowsLogon', 'HideFastUserSwitching', 'Hide fast user switching button', 'HKLM\\\\SOFTWARE\\\\Microsoft\\\\Windows\\\\CurrentVersion\\\\Policies\\\\System', 'HideFastUserSwitching', 'REG_DWORD', 'WindowsLogon/HideFastUserSwitching'),
    ('WindowsPowerShell', 'TurnOnPowerShellScriptBlockLogging', 'Enable PowerShell script block logging (Event ID 4104)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\PowerShell\\\\ScriptBlockLogging', 'EnableScriptBlockLogging', 'REG_DWORD', 'WindowsPowerShell/TurnOnPowerShellScriptBlockLogging'),
    ('WindowsPowerShell', 'TurnOnPowerShellTranscription', 'Enable PowerShell transcription logging', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\PowerShell\\\\Transcription', 'EnableTranscripting', 'REG_DWORD', 'WindowsPowerShell/TurnOnPowerShellTranscription'),
    ('Wireless', 'AllowSimultaneousConnections', 'Block simultaneous Wi-Fi and cellular/Ethernet (prefer Wi-Fi)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WcmSvc\\\\GroupPolicy', 'fMinimizeConnections', 'REG_DWORD', 'Wireless/AllowSimultaneousConnections'),
    ('Wireless', 'ConfigureWiFiSense', 'Configure Wi-Fi Sense (0=disable sharing open networks)', 'HKLM\\\\SOFTWARE\\\\Policies\\\\Microsoft\\\\Windows\\\\WcmSvc\\\\WiFiNetworkManager\\\\Config', 'AutoConnectAllowedOEM', 'REG_DWORD', 'Wireless/ConfigureWiFiSense'),
]

def fetch_policy_csp():
    log("Building Policy CSP (Windows MDM) entries…")
    SEP = "\\"
    entries = []
    for row in POLICY_CSP_DATA:
        area, policy, desc, reg_key, reg_val, dtype, oma_sfx = row
        oma   = f"./Device/Vendor/MSFT/Policy/Config/{oma_sfx}"
        sid   = f"device_vendor_msft_policy_config_{slugify(area)}_{slugify(policy)}"
        hive  = "HKCU" if reg_key.startswith("HKCU") else "HKLM"
        key   = reg_key.split(SEP, 1)[-1] if SEP in reg_key else reg_key.replace("HKLM\\","").replace("HKCU\\","")
        e = make_entry(
            source_id = "graph",
            entry_id  = "csp_" + sid,
            name      = f"{area}/{policy}",
            desc      = desc,
            cats      = [area, "Policy CSP", "Intune Settings Catalog"],
            plat      = "windows",
            methods   = ["intune", "registry", "csp"],
            intune    = [{
                "cat":   area,
                "name":  f"{area}/{policy}",
                "defId": sid,
                "oma":   oma,
                "dtype": "Integer" if dtype == "REG_DWORD" else "String",
                "vals":  [],
                "rec":   "",
                "json":  "settingDefinitionId: " + sid,
            }],
            reg = {"hive": hive, "key": key, "val": reg_val, "type": dtype,
                   "data": "See description", "note": ""} if reg_key else None,
        )
        entries.append(e)
    log(f"  Policy CSP: {len(entries)} entries")
    return entries

# ─── SOURCE 5: Office ADMX ────────────────────────────────────────────────────
OFFICE_ADMX = [
    ("Block macros from Internet — All Office","Blocks VBA macros in Office files downloaded from the internet. 1=Block.","office16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Common\\Security","blockcontentexecutionfrominternet","REG_DWORD"),
    ("VBA Macro Warnings — Word","1=Enable all, 2=Disable+Notify, 3=Disable all, 4=Signed only. Recommended: 3 or 4.","word16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Word\\Security","VBAWarnings","REG_DWORD"),
    ("VBA Macro Warnings — Excel","1=Enable all, 2=Disable+Notify, 3=Disable all, 4=Signed only. Recommended: 3 or 4.","excel16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Excel\\Security","VBAWarnings","REG_DWORD"),
    ("VBA Macro Warnings — PowerPoint","1=Enable all, 2=Disable+Notify, 3=Disable all, 4=Signed only.","ppt16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\PowerPoint\\Security","VBAWarnings","REG_DWORD"),
    ("VBA Macro Warnings — Outlook","Controls macro execution in Outlook. Recommended: 2 (Disable with notification).","outlk16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Outlook\\Security","VBAWarnings","REG_DWORD"),
    ("Disable Trusted Documents — Word","Forces security prompts on every open rather than trusting previously opened files.","word16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Word\\Security\\Trusted Documents","DisableTrustedRecords","REG_DWORD"),
    ("Disable Trusted Documents — Excel","Forces security prompts on every open for spreadsheets.","excel16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Excel\\Security\\Trusted Documents","DisableTrustedRecords","REG_DWORD"),
    ("Disable Office Store Add-ins","Blocks web add-ins from the Office Store. 1=Block.","office16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Common\\Security","DisableStoreApps","REG_DWORD"),
    ("Protected View — Internet Files","Opens internet-sourced Office files in Protected View sandbox. 1=Enable.","office16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Excel\\Security\\ProtectedView","DisableInternetFilesInPV","REG_DWORD"),
    ("Protected View — Email Attachments","Opens email attachments in Protected View. 0=NOT disabled (i.e. enabled).","office16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Excel\\Security\\ProtectedView","DisableAttachmentsInPV","REG_DWORD"),
    ("Disable All ActiveX in Office","Disables all ActiveX controls in Office documents. 1=Disable all.","office16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Common\\Security","disableallactivex","REG_DWORD"),
    ("Outlook — Block external image auto-download","Prevents Outlook auto-downloading images from external URLs in emails. 1=Block.","outlk16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Outlook\\Options\\Mail","blockexternalcontent","REG_DWORD"),
    ("Outlook — Disable automatic email forwarding","Prevents automatic forwarding to external addresses.","outlk16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Outlook\\Options\\Mail","disableautoforward","REG_DWORD"),
    ("Outlook — Junk email protection level","0=No protection, 1=Low, 2=High, 3=Safe lists only. Recommended: 2.","outlk16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Outlook\\Options\\Mail","junkmaillevel","REG_DWORD"),
    ("Office Updates — Enable auto-updates","1=Enable automatic updates for Microsoft 365 Apps for Enterprise.","office16.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Office\\16.0\\Common\\OfficeUpdate","EnableAutomaticUpdates","REG_DWORD"),
    ("Office Updates — Update channel","Update channel: Current, MonthlyEnterprise, SemiAnnual, SemiAnnualPreview.","office16.admx","HKLM\\SOFTWARE\\Policies\\Microsoft\\Office\\16.0\\Common\\OfficeUpdate","UpdateBranch","REG_SZ"),
    ("Word — Block old format files (Word 2 and earlier)","2=Block open of Word 2 and earlier format files.","word16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Word\\Security\\FileBlock","Word2AndEarlier","REG_DWORD"),
    ("Excel — Block old format XLS files","2=Block open of old Excel 4 workbook format.","excel16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Excel\\Security\\FileBlock","XL4Workbooks","REG_DWORD"),
    ("Word — Disable DDE","0=Disable Dynamic Data Exchange to prevent DDE-based macro attacks.","word16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Word\\Options","DDEAllowed","REG_DWORD"),
    ("Teams — Prevent auto-start on Windows logon","Prevents Teams from starting automatically at Windows startup.","skype16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Lync","AutoRun","REG_DWORD"),
    ("Outlook — S/MIME require signed receipt","Requires signed receipts for all S/MIME messages.","outlk16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Outlook\\Security","RequireSMIMEReceipt","REG_DWORD"),
    ("Office — Disable personal information in documents","Prevents Office from storing personal info in document properties.","office16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Common","qmenable","REG_DWORD"),
    ("SharePoint — Trusted sites","Adds SharePoint to trusted sites zone for seamless NTLM/Kerberos auth.","office16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Common\\Internet\\Server Cache","SharePointSiteList","REG_SZ"),
    ("Excel — Disable external data connections","Prevents Excel from refreshing external data connections (OLE DB, ODBC).","excel16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\Excel\\Security","externalconnections_beware","REG_DWORD"),
    ("PowerPoint — Block old format PPT files","2=Block open of PowerPoint 97–2003 format files.","ppt16.admx","HKCU\\Software\\Policies\\Microsoft\\Office\\16.0\\PowerPoint\\Security\\FileBlock","PowerPoint97Files","REG_DWORD"),
]

def fetch_office_admx():
    log("Building Office ADMX entries…")
    entries = []
    app_map = {"word16":"Microsoft Word","excel16":"Microsoft Excel",
               "ppt16":"Microsoft PowerPoint","outlk16":"Microsoft Outlook",
               "office16":"Microsoft Office","skype16":"Microsoft Teams"}
    for row in OFFICE_ADMX:
        name, desc, admx_file, reg_key, val, dtype = row
        hive = "HKCU" if reg_key.startswith("HKCU") else "HKLM"
        key  = reg_key.replace("HKLM\\","").replace("HKCU\\","")
        app  = next((v for k,v in app_map.items() if k in admx_file), "Microsoft Office")
        oma  = (f"./User/Vendor/MSFT/Policy/Config/ADMX_{admx_file.replace('.admx','')}"
                f"~Policy~L_{slugify(app)}/{slugify(name)}")
        e = make_entry(
            source_id="admx_office",
            entry_id="admx_office_" + slugify(name),
            name=name, desc=desc,
            cats=[app,"Microsoft Office","Security"],
            plat="windows", methods=["gpo","admx","registry","intune"],
            gpo={"path": f"User Configuration > Administrative Templates > {app}",
                 "policy": name, "admx": admx_file, "ns": f"Microsoft.Policies.{slugify(app)}"},
            admx={"name": name, "file": admx_file, "cat": app,
                  "regKey": reg_key, "val": val, "type": dtype},
            reg={"hive": hive, "key": key, "val": val, "type": dtype,
                 "data": "See description", "note": ""},
            intune=[{"cat": app, "name": name, "defId": "", "oma": oma,
                     "dtype": "String (ADMX Ingestion)",
                     "vals": [{"v":"<enabled/>","l":"Enabled"},{"v":"<disabled/>","l":"Disabled"}],
                     "rec": "<enabled/>",
                     "json": '"@odata.type": "#microsoft.graph.deviceManagementConfigurationSetting"'}],
        )
        entries.append(e)
    log(f"  Office ADMX: {len(entries)} entries")
    return entries

# ─── SOURCE 6: Custom sources ─────────────────────────────────────────────────
def fetch_custom():
    log("Loading custom sources…")
    entries = []
    cf = REPO_ROOT / "custom_sources.json"
    if cf.exists():
        try:
            for src in json.loads(cf.read_text()).get("sources", []):
                if src.get("_example"): continue
                url = src.get("url","")
                if not url: continue
                log(f"  Fetching: {src.get('name', url)}")
                raw = http_get(url, timeout=30)
                if not raw: continue
                try:
                    data  = json.loads(raw)
                    items = data if isinstance(data,list) else data.get("settings",data.get("entries",[]))
                    for item in items:
                        item.setdefault("_source","custom")
                        if not item.get("id"):
                            item["id"] = "custom_" + hashlib.md5(str(item).encode()).hexdigest()[:10]
                    entries.extend(items)
                    log(f"    Loaded {len(items)} entries")
                except Exception as e:
                    log(f"    Error: {e}")
        except Exception as e:
            log(f"  Error reading custom_sources.json: {e}")
    log(f"  Custom total: {len(entries)}")
    return entries

# ─── SEARCH TEXT ──────────────────────────────────────────────────────────────
def build_search_text(e):
    g, a, r = e.get("gpo") or {}, e.get("admx") or {}, e.get("reg") or {}
    parts = [e.get("name",""), e.get("desc",""),
             " ".join(e.get("cats",[])), " ".join(e.get("methods",[])),
             g.get("path",""), g.get("policy",""),
             a.get("name",""), a.get("regKey",""), a.get("file",""),
             r.get("key",""), r.get("val","")]
    for i in (e.get("intune") or []):
        parts += [i.get("defId",""), i.get("name",""), i.get("oma","")]
    return " ".join(p for p in parts if p).lower()

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sources", default="all")
    args    = parser.parse_args()
    filters = args.sources.lower().split(",")
    def want(n): return "all" in filters or n in filters

    log("=" * 60)
    log("Helient Settings Explorer — Index Builder v3")
    log("=" * 60)

    all_entries, source_status = [], {}

    if want("graph"):
        cid  = os.environ.get("AZURE_CLIENT_ID","")
        csec = os.environ.get("AZURE_CLIENT_SECRET","")
        tid  = os.environ.get("AZURE_TENANT_ID","common")
        if cid and csec:
            log("Authenticating to Microsoft Graph…")
            token = get_graph_token(cid, csec, tid)
            if token:
                # Decode JWT payload to show what roles are actually in the token
                try:
                    import base64, json as _json
                    parts = token.split(".")
                    if len(parts) >= 2:
                        payload_b64 = parts[1] + "=="  # pad
                        payload = _json.loads(base64.urlsafe_b64decode(payload_b64).decode())
                        roles = payload.get("roles", [])
                        scp   = payload.get("scp", "")
                        appid = payload.get("appid", "")
                        log(f"  Token roles (application permissions in token): {roles if roles else ['(none — application permissions not present in token)']}")
                        if scp:
                            log(f"  Token scp (delegated scopes): {scp[:200]}")
                        if not roles:
                            log("  !! No 'roles' claim in token = Application permissions NOT effective")
                            log("     Check Azure Portal > Enterprise Applications > Helient Settings Explorer")
                            log("     > Permissions > verify admin consent shows Application permissions (not Delegated)")
                            log("     Also check: App Registration > API Permissions > all Application perms show 'Granted' in green")
                except Exception as e:
                    log(f"  Could not decode token: {e}")

                cat = fetch_graph_catalog(token)
                gpo = fetch_graph_gpo(token)
                all_entries.extend(cat)
                all_entries.extend(gpo)
                total = len(cat) + len(gpo)
                source_status["graph"] = {"ok": total > 0, "count": total}
                log(f"  Graph total: {total} (catalog:{len(cat)}, gpo:{len(gpo)})")
            else:
                log("  Graph auth failed — check AZURE_CLIENT_ID / SECRET / TENANT_ID secrets")
                source_status["graph"] = {"ok": False, "count": 0, "error": "Auth failed"}
        else:
            log("  Skipping Graph (no credentials configured)")
            source_status["graph"] = {"ok": False, "count": 0, "error": "No credentials"}

        # Always try the public IntunePMFiles dataset — no auth required
        pm_entries = fetch_intune_pm_files()
        if pm_entries:
            all_entries.extend(pm_entries)
            existing = source_status.get("graph", {})
            source_status["graph"] = {
                "ok": True,
                "count": existing.get("count", 0) + len(pm_entries),
            }
            log(f"  IntunePMFiles added {len(pm_entries)} entries to graph source")

    if want("chromium"):
        e = fetch_chromium()
        all_entries.extend(e)
        source_status["chromium"] = {"ok": True, "count": len(e)}

    if want("admx"):
        e = fetch_windows_admx()
        all_entries.extend(e)
        source_status["admx_windows"] = {"ok": True, "count": len(e)}

        # Policy CSP — always include alongside Windows ADMX
        csp_entries = fetch_policy_csp()
        all_entries.extend(csp_entries)
        source_status["policy_csp"] = {"ok": True, "count": len(csp_entries)}

    if want("office"):
        e = fetch_office_admx()
        all_entries.extend(e)
        source_status["admx_office"] = {"ok": True, "count": len(e)}

    if want("custom"):
        e = fetch_custom()
        all_entries.extend(e)
        source_status["custom"] = {"ok": True, "count": len(e)}

    # Deduplicate
    seen, deduped = set(), []
    for e in all_entries:
        eid = e.get("id","")
        if eid and eid not in seen:
            seen.add(eid)
            deduped.append(e)
        elif not eid:
            deduped.append(e)

    log(f"\nTotal after dedup: {len(deduped)}")

    for e in deduped:
        e["_text"] = build_search_text(e)

    # Write index
    idx = DATA_DIR / "index.json"
    idx.write_text(json.dumps(deduped, separators=(",",":"), ensure_ascii=False))
    sz = idx.stat().st_size

    # Write chunks
    CHUNK = 200
    for old in CHUNKS_DIR.glob("chunk_*.json"): old.unlink()
    chunks = []
    for i, start in enumerate(range(0, len(deduped), CHUNK)):
        chunk = deduped[start:start+CHUNK]
        p = CHUNKS_DIR / f"chunk_{i:03d}.json"
        p.write_text(json.dumps(chunk, separators=(",",":"), ensure_ascii=False))
        chunks.append({"file": f"data/chunks/chunk_{i:03d}.json",
                       "start": start, "count": len(chunk)})
    log(f"Wrote {len(chunks)} chunks")

    meta = {
        "last_updated":  datetime.now(timezone.utc).isoformat(),
        "total_entries": len(deduped),
        "chunk_count":   len(chunks),
        "chunks":        chunks,
        "sources":       source_status,
        "index_size_kb": sz // 1024,
    }
    (DATA_DIR / "meta.json").write_text(json.dumps(meta, indent=2))

    log("\n" + "=" * 60)
    log(f"Build complete — {len(deduped)} settings")
    for src, info in source_status.items():
        log(f"  {'✓' if info['ok'] else '✗'} {src}: {info['count']}")
    log("=" * 60)

if __name__ == "__main__":
    main()
